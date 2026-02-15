import sys
from datetime import datetime, timedelta
import json
import os
import platform
import subprocess
import hashlib  # For password hashing
import random  # For OTP generation
import time    # For simulating OTP delay

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QFrame, QMessageBox, QSizePolicy, QPushButton, QScrollArea, QGridLayout,
    QLineEdit, QGroupBox, QInputDialog, QTableWidget, QTableWidgetItem,
    QDialog, QHeaderView, QComboBox, QAbstractItemView, QFileDialog, QMenu,
    QDateEdit, QFormLayout, QStackedWidget
)
from PyQt6.QtGui import QFont, QPixmap, QImageReader, QAction
from PyQt6.QtCore import Qt, QDate, QTimer

# ======== DB SECTION =========
import sqlite3
import os
import hashlib
from typing import List, Dict, Any, Self

def _safefloat(v):
    try:
        return float(str(v).replace(",", "").strip() or 0)
    except Exception:
        return 0.0

class DBManager:
    def __init__(self, dbpath=None):
        if dbpath is None:
            # Detect if running as PyInstaller executable
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
            else:
                base_dir = os.path.dirname(__file__)

            dbpath = os.path.join(base_dir, "ktsdatabase.db")

        self.dbpath = dbpath
        new_db = not os.path.exists(dbpath)
        self.conn = sqlite3.connect(self.dbpath, check_same_thread=False)

        self.conn.row_factory = sqlite3.Row
        self._configure()
        if new_db:
            self.create_tables()

    def _configure(self):
        c = self.conn.cursor()
        c.execute("PRAGMA foreign_keys = ON;")
        c.execute("PRAGMA journal_mode = WAL;")
        c.execute("PRAGMA synchronous = NORMAL;")
        self.conn.commit()

    def create_tables(self):
        c = self.conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS trips (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            date            TEXT,
            vehicle_no      TEXT,
            location_from_to TEXT,
            broker_office   TEXT,
            driver_amount   REAL,
            profit          REAL,
            expense         REAL,
            total           REAL,
            status          TEXT,
            detail_json     TEXT
        )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS vehicle_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                vehicle_no TEXT,
                fc_expense REAL,
                tyre_amount REAL,
                tyre_type TEXT,
                tax REAL,
                tax_type TEXT,
                spare_work REAL,
                spare_type TEXT,
                loan REAL,
                insurance REAL,
                others REAL,
                remarks TEXT,
                total REAL
            )
        ''')
        # ---- FIXED OFFICE EXPENSES TABLE ----
        c.execute('''
            CREATE TABLE IF NOT EXISTS office_expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month TEXT,
                current_bill REAL,
                manager_salary REAL,
                office_rent REAL,
                others REAL,
                total REAL
            )
        ''')
        # New table for Vehicle and Driver Details
        c.execute('''
            CREATE TABLE IF NOT EXISTS vehicle_driver_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vehicle_no TEXT UNIQUE,
                registration_date TEXT,
                fitness_upto TEXT,
                tax_upto TEXT,
                insurance_upto TEXT,
                pucc_upto TEXT,
                permit_upto TEXT,
                national_permit_upto TEXT,
                driver_name TEXT,
                driver_contact TEXT,
                driver_alt_contact TEXT,
                driver_experience TEXT,
                driver_adhar TEXT,
                driver_license_path TEXT,
                loan_total REAL,
                loan_paid REAL,
                loan_remaining REAL,
                driver_date_of_joining TEXT,
                driver_bank_account TEXT
            )
        ''')
        # UPDATED: User table for authentication with permanent and additional mobile
        c.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    permanent_mobile TEXT,
                    additional_mobile TEXT
            )
        ''')
        self.conn.commit()

        # UPDATED: Add a default user if none exist
        self._add_default_user_if_needed()

    
    def _add_default_user_if_needed(self):
        c = self.conn.cursor()
        c.execute("SELECT COUNT(*) FROM users")
        if c.fetchone()[0] == 0:
            hashed = self.hash_password("password")
            c.execute(
                "INSERT INTO users (username, password_hash, permanent_mobile, additional_mobile) VALUES (?, ?, ?, ?)",
                ("admin", hashed, "+91 8637484827", "")
            )
            self.conn.commit()
            print("Default user created: admin / password")

    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()

    def verify_login(self, username, password):
        c = self.conn.cursor()
        c.execute("SELECT password_hash FROM users WHERE username=?", (username,))
        row = c.fetchone()
        if not row:
            return False
        return self.hash_password(password) == row["password_hash"]

    def update_password(self, username, new_password):
        c = self.conn.cursor()
        hashed = self.hash_password(new_password)
        c.execute("UPDATE users SET password_hash=? WHERE username=?", (hashed, username))
        self.conn.commit()
        return c.rowcount

    def reset_password(self):
        newpass = self.newpasswordinput.text()
        confirmpass = self.confirmpasswordinput.text()
        if not newpass or not confirmpass:
            QMessageBox.warning(self, "Input Error", "Please enter and confirm your new password.")
            return
        if newpass != confirmpass:
            QMessageBox.warning(self, "Mismatch", "New password and confirmation do not match.")
            return
        if len(newpass) < 6:
            QMessageBox.warning(self, "Weak Password", "Password must be at least 6 characters.")
            return

        result = self.db.updatepassword(self.currentusername, newpass)
        if result > 0:
            QMessageBox.information(self, "Success", f"Password reset successfully for user {self.currentusername}")
            self.accept()
        else:
            QMessageBox.warning(self, "Error", "Failed to reset password. User may not exist.")

    # ---------------- TRIPS ----------------

    def load_trips(self) -> List[sqlite3.Row]:
        return self.conn.execute("SELECT * FROM trips ORDER BY date ASC").fetchall()

    def loadtrips(self):  # legacy name alias
        return self.load_trips()

    def save_trip(self, data: tuple) -> int:
        cur = self.conn.cursor()
        if len(data) == 9:   # legacy call – no detail_json
            cur.execute('''
                INSERT INTO trips
                (date, vehicle_no, location_from_to, broker_office,
                 driver_amount, profit, expense, total, status)
                VALUES (?,?,?,?,?,?,?,?,?)
            ''', data)
        elif len(data) == 10:            # new call – with detail_json
            cur.execute('''
                INSERT INTO trips
                (date, vehicle_no, location_from_to, broker_office,
                 driver_amount, profit, expense, total, status, detail_json)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            ''', data)
        else:
            raise ValueError("save_trip expected 9 or 10 fields")
        self.conn.commit()
        return cur.lastrowid

    def savetrip(self, data: tuple) -> int:
        return self.save_trip(data)

    def update_trip(self, tripid: int, data: tuple):
        if len(data) == 9:
            sql = '''
                UPDATE trips SET
                date=?, vehicle_no=?, location_from_to=?, broker_office=?,
                driver_amount=?, profit=?, expense=?, total=?, status=?
                WHERE id=?
            '''
        elif len(data) == 10:
            sql = '''
                UPDATE trips SET
                date=?, vehicle_no=?, location_from_to=?, broker_office=?,
                driver_amount=?, profit=?, expense=?, total=?, status=?,
                detail_json=?
                WHERE id=?
            '''
        else:
            raise ValueError("update_trip expected 9 or 10 fields")
        self.conn.execute(sql, (*data, tripid))
        self.conn.commit()

    def updatetrip(self, tripid: int, data: tuple):
        return self.update_trip(tripid, data)

    def delete_trip(self, trip_id: int):
        """Delete a trip by its id."""
        self.conn.execute("DELETE FROM trips WHERE id=?", (trip_id,))
        self.conn.commit()

    def delete_office_expense(self, expid: int):
        self.conn.execute("DELETE FROM office_expenses WHERE id=?", (expid,))
        self.conn.commit()

    def deleteofficeexpense(self, expid):
        return self.delete_office_expense(expid)

    # ---------------- VEHICLE EXPENSES ----------------
    def load_vehicle_expenses(self) -> List[sqlite3.Row]:
        return self.conn.execute("SELECT * FROM vehicle_expenses ORDER BY date DESC").fetchall()

    def loadvehicleexpenses(self):
        return self.load_vehicle_expenses()

    def save_vehicle_expense(self, data: List[Any]) -> int:
        """
        Accepts either:
        - 13 fields (date, vehicle_no, fc_expense, tyre_amount, tyre_type, tax, tax_type, spare_work, spare_type, loan, insurance, others, remarks)
          => total computed automatically
        - or 14 fields (same plus total)
        """
        d = list(data)
        if len(d) == 13:
            total = sum(_safefloat(d[i]) for i in (2, 3, 5, 7, 9, 10, 11))
            d.append(total)
        cur = self.conn.cursor()
        cur.execute('''
            INSERT INTO vehicle_expenses (date,vehicle_no,fc_expense,tyre_amount,tyre_type,tax,tax_type,spare_work,spare_type,loan,insurance,others,remarks,total)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', tuple(d))
        self.conn.commit()
        return cur.lastrowid

    def savevehicleexpense(self, data):
        return self.save_vehicle_expense(data)

    def update_vehicle_expense(self, expid: int, data: List[Any]):
        d = list(data)
        if len(d) == 13:
            total = sum(_safefloat(d[i]) for i in (2, 3, 5, 7, 9, 10, 11))
            d.append(total)
        self.conn.execute('''
            UPDATE vehicle_expenses SET
                date=?, vehicle_no=?, fc_expense=?, tyre_amount=?, tyre_type=?, tax=?, tax_type=?,
                spare_work=?, spare_type=?, loan=?, insurance=?, others=?, remarks=?, total=?
            WHERE id=?
        ''', tuple(d) + (expid,))
        self.conn.commit()

    def updatevehicleexpense(self, expid, data):
        return self.update_vehicle_expense(expid, data)

    def delete_vehicle_expense(self, expid: int):
        self.conn.execute("DELETE FROM vehicle_expenses WHERE id=?", (expid,))
        self.conn.commit()

    def deletevehicleexpense(self, expid):
        return self.delete_vehicle_expense(expid)

    # ---------------- OFFICE EXPENSES ----------------
    def load_office_expenses(self) -> List[sqlite3.Row]:
        return self.conn.execute("SELECT * FROM office_expenses ORDER BY month DESC").fetchall()

    def loadofficeexpenses(self):
        return self.load_office_expenses()

    def save_office_expense(self, data: tuple) -> int:
        """
        data: (month, current_bill, manager_salary, office_rent, others)
        """
        total = sum(_safefloat(x) for x in data[1:])  # Automatically calculate total
        cur = self.conn.cursor()
        cur.execute('''
            INSERT INTO office_expenses (month, current_bill, manager_salary, office_rent, others, total)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (*data, total))
        self.conn.commit()
        return cur.lastrowid


    def saveofficeexpense(self, data):
        return self.save_office_expense(data)

    def update_office_expense(self, expid: int, data: tuple):
        total = sum(_safefloat(x) for x in data[1:])  # Recalculate total
        self.conn.execute('''
            UPDATE office_expenses
            SET month=?, current_bill=?, manager_salary=?, office_rent=?, others=?, total=?
            WHERE id=?
        ''', (*data, total, expid))
        self.conn.commit()

    def updateofficeexpense(self, expid, data):
        return self.update_office_expense(expid, data)

    # ---------------- VEHICLE DRIVER DETAILS ----------------
    def save_vehicle_driver_details(self, data: Dict[str, Any]) -> int:
        cur = self.conn.cursor()
        # Check if vehicle_no already exists
        existing_vehicle = self.conn.execute(
            "SELECT id FROM vehicle_driver_details WHERE vehicle_no = ?", (data['vehicle_no'],)
        ).fetchone()

        if existing_vehicle:
            # Update existing record
            self.update_vehicle_driver_details(existing_vehicle['id'], data)
            return existing_vehicle['id']
        else:
            # Insert new record
            cur.execute('''
                INSERT INTO vehicle_driver_details (
                    vehicle_no, registration_date, fitness_upto, tax_upto, insurance_upto,
                    pucc_upto, permit_upto, national_permit_upto, driver_name, driver_contact,
                    driver_alt_contact, driver_experience, driver_adhar, driver_license_path,
                    loan_total, loan_paid, loan_remaining, driver_date_of_joining, driver_bank_account
                ) VALUES (
                    :vehicle_no, :registration_date, :fitness_upto, :tax_upto, :insurance_upto,
                    :pucc_upto, :permit_upto, :national_permit_upto, :driver_name, :driver_contact,
                    :driver_alt_contact, :driver_experience, :driver_adhar, :driver_license_path,
                    :loan_total, :loan_paid, :loan_remaining, :driver_date_of_joining, :driver_bank_account
                )
            ''', data)
            self.conn.commit()
            return cur.lastrowid

    def update_vehicle_driver_details(self, record_id: int, data: Dict[str, Any]):
        self.conn.execute('''
            UPDATE vehicle_driver_details SET
                vehicle_no = :vehicle_no,
                registration_date = :registration_date,
                fitness_upto = :fitness_upto,
                tax_upto = :tax_upto,
                insurance_upto = :insurance_upto,
                pucc_upto = :pucc_upto,
                permit_upto = :permit_upto,
                national_permit_upto = :national_permit_upto,
                driver_name = :driver_name,
                driver_contact = :driver_contact,
                driver_alt_contact = :driver_alt_contact,
                driver_experience = :driver_experience,
                driver_adhar = :driver_adhar,
                driver_license_path = :driver_license_path,
                loan_total = :loan_total,
                loan_paid = :loan_paid,
                loan_remaining = :loan_remaining,
                driver_date_of_joining = :driver_date_of_joining,
                driver_bank_account = :driver_bank_account
            WHERE id = :id
        ''', {**data, 'id': record_id})
        self.conn.commit()

    def load_vehicle_driver_details(self, vehicle_no: str) -> Dict[str, Any] | None:
        row = self.conn.execute(
            "SELECT * FROM vehicle_driver_details WHERE vehicle_no = ?", (vehicle_no,)
        ).fetchone()
        return dict(row) if row else None

    def load_all_vehicle_driver_details(self) -> List[Dict[str, Any]]:
        rows = self.conn.execute("SELECT * FROM vehicle_driver_details").fetchall()
        return [dict(row) for row in rows]

    def delete_vehicle_driver_details(self, vehicle_no: str):
        self.conn.execute("DELETE FROM vehicle_driver_details WHERE vehicle_no = ?", (vehicle_no,))
        self.conn.commit()

    # ---------------- Helpers ----------------
    def row_to_dict(self, row: sqlite3.Row) -> Dict[str, Any]:
        return dict(row) if row is not None else {}

    def trips_as_list(self) -> List[Dict[str, Any]]:
        return [self.row_to_dict(r) for r in self.load_trips()]

    def vehicle_expenses_as_list(self) -> List[Dict[str, Any]]:
        return [self.row_to_dict(r) for r in self.load_vehicle_expenses()]

    def office_expenses_as_list(self) -> List[Dict[str, Any]]:
        return [self.row_to_dict(r) for r in self.load_office_expenses()]

# ======== END OF DB SECTION ==================================================


# Optional PDF / Excel export dependencies
try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    import pandas as pd
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("openpyxl not found – Excel export will use pandas.")
except ImportError:
    print("ReportLab/pandas not found – PDF & Excel export disabled.")


# ----------------------------------------------------------------------
# Utility helpers
# ----------------------------------------------------------------------


NAVY = "#2c3e50"
NAVY_DARK = "#34495e"
WHITE = "#ffffff"
LIGHT_GRAY = "#f7f7f7"
SELECT_ROW_COLOR = "#dcdcdc"


def safe_float(value: str | float, default=0.0) -> float:
    try:
        return float(str(value).replace(",", "").strip() or "0")
    except Exception:
        return default


def show_int_amount(value: str | float) -> str:
    try:
        return f"{int(round(safe_float(value))):,}"
    except Exception:
        return "0"



def parse_date_from_display(date_str):
    """Parse date from dd-mm-yyyy format to datetime object"""
    try:
        return datetime.strptime(date_str, "%d-%m-%Y").date()
    except:
        try:
            # Fallback for yyyy-mm-dd format (for backward compatibility)
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except:
            return None

def format_date_for_display(date_obj):
    """Format date object to dd-mm-yyyy string"""
    if isinstance(date_obj, str):
        # If it's already a string, try to parse and reformat
        parsed_date = parse_date_from_display(date_obj)
        if parsed_date:
            return parsed_date.strftime("%d-%m-%Y")
        return date_obj
    elif hasattr(date_obj, 'strftime'):
        return date_obj.strftime("%d-%m-%Y")
    else:
        return str(date_obj)

# ----------------------------------------------------------------------
# Global stylesheets
# ----------------------------------------------------------------------


BLUE_WHITE_STYLESHEET = """
QMainWindow, QWidget, QDialog {
    background-color: #f0f5fa;
    color: #1a237e;
    font-family: Arial, sans-serif;
    font-size: 14px;
}
QLabel { color: #1a237e; font-weight: 600; }
QLineEdit, QDateEdit, QComboBox {
    background-color: #ffffff;
    border: 2px solid #1a237e;
    border-radius: 5px;
    padding: 4px 8px;
    color: #1a237e;
}
QPushButton {
    background-color: #1a237e; color: #ffffff;
    border-radius: 6px; font-weight: bold;
    min-width: 60px; min-height: 30px;
}
QPushButton:hover { background-color: #3f51b5; }
QTableWidget {
    background-color: #ffffff;
    gridline-color: #90caf9;
    color: #1a237e; border: none;
    font-size: 14px;
}
QHeaderView::section {
    background-color: #1a237e;
    color: white; font-weight: bold; font-size: 14px;
}
"""

MAIN_STYLESHEET = """
QWidget { background-color: #f0f4f8; color: #003366; }
QScrollArea { border: none; background-color: #f0f4f8; }
"""


# ======================================================================
# NEW: Authentication Dialogs
# ======================================================================


class LoginDialog(QDialog):
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.setWindowTitle("Login")
        self.setFixedSize(400, 350)
        self.setStyleSheet(BLUE_WHITE_STYLESHEET)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        title = QLabel("KTS Transport Login")
        title.setFont(QFont("Arial Black", 18))
        layout.addWidget(title, alignment=Qt.AlignmentFlag.AlignCenter)

        form = QFormLayout()
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Username")

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Password")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)

        form.addRow("Username:", self.username_input)
        form.addRow("Password:", self.password_input)
        layout.addLayout(form)

        login_btn = QPushButton("Login")
        login_btn.clicked.connect(self.attempt_login)
        layout.addWidget(login_btn)

        forgot_btn = QPushButton("Forgot Password?")
        forgot_btn.setStyleSheet("background:transparent;color:#1a237e;border:none;")
        forgot_btn.clicked.connect(self.show_forgot_password_dialog)
        layout.addWidget(forgot_btn, alignment=Qt.AlignmentFlag.AlignCenter)

    def attempt_login(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        if self.db.verify_login(username, password):
            self.accept()
        else:
            QMessageBox.warning(self, "Login Failed", "Invalid username or password.")
            self.password_input.clear()

    def show_forgot_password_dialog(self):
        dlg = ForgotPasswordDialog(self.db, self)
        dlg.exec()


# Email OTP using Gmail SMTP and app password

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

class EmailOTP:
    def __init__(self):
        self.otp = None

    def send_otp(self, receiver_email):  # ✅ must accept receiver_email
        sender_email = "p.sathishkts@gmail.com"  # Replace with your Gmail
        sender_password = "mslb wfth dkjp dozq"   # Your Gmail app password

        self.otp = str(random.randint(100000, 999999))

        subject = "Your OTP Code"
        body = f"Your One-Time Password (OTP) is: {self.otp}\nThis code will expire in 5 minutes."
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = receiver_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        try:
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
            server.quit()
            QMessageBox.information(None, "OTP Sent", f"OTP sent to {receiver_email}")
        except Exception as e:
            QMessageBox.warning(None, "Error", f"Failed to send email: {e}")


# Main dialog for reset password flow with email OTP

class ForgotPasswordDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Forgot Password (OTP)")
        self.setFixedSize(500, 270)
        self.setStyleSheet(BLUE_WHITE_STYLESHEET)
        self.emailotp = EmailOTP()
        self.generatedotp = None
        self.otpexpiry = 0
        self.username = ""
        
        layout = QVBoxLayout(self)
        layout.setSpacing(25)

        # Username row with Send OTP button
        username_row = QHBoxLayout()
        username_label = QLabel("Username:")
        username_label.setMinimumWidth(150)
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Enter Username")
        
        self.send_otp_btn = QPushButton("Send OTP")
        self.send_otp_btn.setMinimumWidth(130)
        self.send_otp_btn.clicked.connect(self.send_otp)
        
        username_row.addWidget(username_label)
        username_row.addWidget(self.username_input, 1)  # stretch factor 1
        username_row.addWidget(self.send_otp_btn)
        layout.addLayout(username_row)

        # OTP row with Verify button
        otp_row = QHBoxLayout()
        otp_label = QLabel("Enter OTP:")
        otp_label.setMinimumWidth(150)
        self.otp_input = QLineEdit()
        self.otp_input.setPlaceholderText("Enter OTP")
        self.otp_input.setEnabled(False)
        
        self.verify_otp_btn = QPushButton("Verify OTP")
        self.verify_otp_btn.setMinimumWidth(130)
        self.verify_otp_btn.setEnabled(False)
        self.verify_otp_btn.clicked.connect(self.verify_otp)
        
        otp_row.addWidget(otp_label)
        otp_row.addWidget(self.otp_input, 1)  # stretch factor 1
        otp_row.addWidget(self.verify_otp_btn)
        layout.addLayout(otp_row)

        # New Password row (no button)
        new_pass_row = QHBoxLayout()
        new_pass_label = QLabel("New Password:")
        new_pass_label.setMinimumWidth(150)
        self.new_pass_input = QLineEdit()
        self.new_pass_input.setPlaceholderText("New Password")
        self.new_pass_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.new_pass_input.setEnabled(False)
        
        new_pass_row.addWidget(new_pass_label)
        new_pass_row.addWidget(self.new_pass_input, 1)
        layout.addLayout(new_pass_row)

        # Confirm Password row (no button)
        confirm_pass_row = QHBoxLayout()
        confirm_pass_label = QLabel("Confirm Password:")
        confirm_pass_label.setMinimumWidth(150)
        self.confirm_pass_input = QLineEdit()
        self.confirm_pass_input.setPlaceholderText("Confirm Password")
        self.confirm_pass_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.confirm_pass_input.setEnabled(False)
        
        confirm_pass_row.addWidget(confirm_pass_label)
        confirm_pass_row.addWidget(self.confirm_pass_input, 1)
        layout.addLayout(confirm_pass_row)

        # Reset Password button (full width at bottom)
        self.reset_btn = QPushButton("Reset Password")
        self.reset_btn.setEnabled(False)
        self.reset_btn.setMinimumHeight(40)
        self.reset_btn.clicked.connect(self.reset_password)
        layout.addWidget(self.reset_btn)
        
        layout.addStretch()  # Push everything to the top

    def send_otp(self):
        self.username = self.username_input.text().strip()
        if not self.username:
            QMessageBox.warning(self, "Error", "Enter username.")
            return
        # Check for username existence in DB
        c = self.db.conn.cursor()
        c.execute("SELECT id FROM users WHERE username=?", (self.username,))
        if not c.fetchone():
            QMessageBox.warning(self, "Error", "Username not found.")
            return
        
        receiver = "p.sathishkts@gmail.com"
        self.emailotp.send_otp(receiver)
        self.generatedotp = self.emailotp.otp
        import time
        self.otpexpiry = time.time() + 300
        self.otp_input.setEnabled(True)
        self.verify_otp_btn.setEnabled(True)

    def verify_otp(self):
        import time
        entered = self.otp_input.text().strip()
        if time.time() > self.otpexpiry:
            QMessageBox.warning(self, "Expired", "OTP expired. Send again.")
            return
        if entered == self.generatedotp:
            QMessageBox.information(self, "Verified", "OTP verified successfully.")
            self.new_pass_input.setEnabled(True)
            self.confirm_pass_input.setEnabled(True)
            self.reset_btn.setEnabled(True)
            self.generatedotp = None
        else:
            QMessageBox.warning(self, "Error", "Invalid OTP.")

    def reset_password(self):
        new = self.new_pass_input.text()
        confirm = self.confirm_pass_input.text()
        if not new or not confirm:
            QMessageBox.warning(self, "Error", "Enter password fields.")
            return
        if new != confirm:
            QMessageBox.warning(self, "Error", "Passwords do not match.")
            return
        if len(new) < 6:
            QMessageBox.warning(self, "Error", "Password too short (min 6).")
            return
        updated = self.db.update_password(self.username, new)
        if updated:
            QMessageBox.information(self, "Success", "Password reset successfully.")
            self.accept()
        else:
            QMessageBox.warning(self, "Error", "Username not found.")


# ======================================================================
#  VEHICLE EXPENSES – Dialog & Page
# ======================================================================



class ExpenseDialog(QDialog):
    """Dialog for adding/editing a single vehicle-expense record."""
    def __init__(self, parent=None, data=None):
        super().__init__(parent)
        self.setWindowTitle("New / Edit Expense")
        self.setMinimumWidth(420)
        self.setStyleSheet(BLUE_WHITE_STYLESHEET)

        layout = QFormLayout(self)

        self.date_edit = QDateEdit(calendarPopup=True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        layout.addRow("Date:", self.date_edit)

        self.vehicle_no_edit = QLineEdit()
        layout.addRow("Vehicle No:", self.vehicle_no_edit)

        def _money_edit():
            e = QLineEdit("0")
            e.setAlignment(Qt.AlignmentFlag.AlignRight)
            return e

        self.fc_expense_edit       = _money_edit()
        self.tyre_amount_edit      = _money_edit()
        self.tax_amount_edit       = _money_edit()
        self.spare_work_amount_edit= _money_edit()
        self.loan_edit             = _money_edit()
        self.insurance_edit        = _money_edit()
        self.others_edit           = _money_edit()

        self.spare_work_type_combo = QComboBox()
        self.spare_work_type_combo.addItems(
            ["None", "Spares", "Body Work", "Trailor Work",
             "Mechanical Work", "Electrical Work"])

        self.tyre_type_combo = QComboBox()
        self.tyre_type_combo.addItems(["None", "New", "Old"])

        self.tax_type_combo = QComboBox()
        self.tax_type_combo.addItems(
            ["None", "Quarter tax", "NP tax", "Green tax", "Five year tax"])

        spare_h = QHBoxLayout(); spare_h.addWidget(self.spare_work_amount_edit); spare_h.addWidget(self.spare_work_type_combo)
        tyre_h  = QHBoxLayout(); tyre_h.addWidget(self.tyre_amount_edit); tyre_h.addWidget(self.tyre_type_combo)
        tax_h   = QHBoxLayout(); tax_h.addWidget(self.tax_amount_edit);  tax_h.addWidget(self.tax_type_combo)

        layout.addRow("FC Expense:", self.fc_expense_edit)
        layout.addRow("Spare Work:", spare_h)
        layout.addRow("Tyre Expense:", tyre_h)
        layout.addRow("Tax:", tax_h)
        layout.addRow("Loan:", self.loan_edit)
        layout.addRow("Insurance:", self.insurance_edit)
        layout.addRow("Others:", self.others_edit)

        self.remarks_edit = QLineEdit()
        layout.addRow("Remarks:", self.remarks_edit)

        btn = QPushButton("Save"); layout.addRow(btn)
        btn.clicked.connect(self.accept)

        if data: self.load_data(data)

    # ------------------------------------------------------------------
    
    def load_data(self, d):
        d += [""] * (13-len(d))
        self.date_edit.setDate(QDate.fromString(d[0], "yyyy-MM-dd"))
        self.vehicle_no_edit.setText(d[1])
        self.fc_expense_edit.setText(d[2]);  self.tyre_amount_edit.setText(d[3])
        self.tyre_type_combo.setCurrentText(d[4])
        self.tax_amount_edit.setText(d[5]); self.tax_type_combo.setCurrentText(d[6])
        self.spare_work_amount_edit.setText(d[7]); self.spare_work_type_combo.setCurrentText(d[8])
        self.loan_edit.setText(d[9]); self.insurance_edit.setText(d[10])
        self.others_edit.setText(d[11]); self.remarks_edit.setText(d[12])

    def data(self):
        return [
            self.date_edit.date().toString("yyyy-MM-dd"),
            self.vehicle_no_edit.text().strip(),
            self.fc_expense_edit.text().strip(),
            self.tyre_amount_edit.text().strip(),
            self.tyre_type_combo.currentText(),
            self.tax_amount_edit.text().strip(),
            self.tax_type_combo.currentText(),
            self.spare_work_amount_edit.text().strip(),
            self.spare_work_type_combo.currentText(),
            self.loan_edit.text().strip(),
            self.insurance_edit.text().strip(),
            self.others_edit.text().strip(),
            self.remarks_edit.text().strip(),
        ]


class VehicleExpensePage(QWidget):
    def __init__(self, back_cb=None, db=None):
        super().__init__()
        self.back_cb = back_cb
        self.db = db
        self.records = []
        self.record_ids = []
        self.initui()
        if self.db:
            self.load_from_db()
        else:
            self.refresh()

    def initui(self):
        vbox = QVBoxLayout(self)
        vbox.setContentsMargins(20, 20, 20, 20)
        vbox.setSpacing(15)
        self.setStyleSheet(MAIN_STYLESHEET)

        # Header
        header_layout = QHBoxLayout()
        header_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label = QLabel("Vehicle Expenses")
        title_label.setFont(QFont("Arial Black", 24))
        title_label.setStyleSheet("color: #003366; font-weight: bold;")
        header_layout.addWidget(title_label)
        vbox.addLayout(header_layout)

        # ------------ toolbar with colored buttons ------------
        tbar = QHBoxLayout()
        
        # Home Button (Orange)
        home_btn = QPushButton("← Home")
        home_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF6B35;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px 16px;
                border-radius: 6px;
                min-width: 100px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #E85A2B;
            }
        """)
        home_btn.clicked.connect(lambda: self.back_cb())
        tbar.addWidget(home_btn)
        
        # NEW Button (Green)
        self.new_btn = QPushButton("NEW")
        self.new_btn.setStyleSheet("""
            QPushButton {
                background-color: #28A745;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px 16px;
                border-radius: 6px;
                min-width: 100px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.new_btn.clicked.connect(self.new_record)
        tbar.addWidget(self.new_btn)
        
        # Vehicle Label and Input
        vehicle_label = QLabel("Vehicle:")
        vehicle_label.setStyleSheet("color: #003366; font-weight: bold; font-size: 14px;")
        tbar.addWidget(vehicle_label)
        
        self.filt_vehicle = QLineEdit()
        self.filt_vehicle.setPlaceholderText("Vehicle No.")
        self.filt_vehicle.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 2px solid #87CEEB;
                border-radius: 5px;
                padding: 6px;
                font-size: 13px;
                min-width: 150px;
            }
        """)
        tbar.addWidget(self.filt_vehicle)
        
        # Date Label and Combo
        date_label = QLabel("Date:")
        date_label.setStyleSheet("color: #003366; font-weight: bold; font-size: 14px;")
        tbar.addWidget(date_label)
        
        self.date_combo = QComboBox()
        self.date_combo.addItems(
            ["All", "Last 1 Day", "Last Month", "Last 3 Months",
             "Last 6 Months", "Last Year", "Custom"])
        self.date_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
                border: 2px solid #87CEEB;
                border-radius: 5px;
                padding: 6px;
                font-size: 13px;
                min-width: 130px;
            }
        """)
        self.date_combo.currentTextChanged.connect(self._date_combo_changed)
        tbar.addWidget(self.date_combo)
        
        # Date range inputs
        self.from_date = QDateEdit(calendarPopup=True)
        self.from_date.setDisplayFormat("yyyy-MM-dd")
        self.from_date.setStyleSheet("""
            QDateEdit {
                background-color: white;
                border: 2px solid #87CEEB;
                border-radius: 5px;
                padding: 6px;
                font-size: 13px;
            }
        """)
        self.from_date.hide()
        tbar.addWidget(self.from_date)
        
        to_label = QLabel("to")
        to_label.setStyleSheet("color: #003366; font-weight: bold;")
        to_label.hide()
        self.to_label = to_label
        tbar.addWidget(to_label)
        
        self.to_date = QDateEdit(calendarPopup=True)
        self.to_date.setDisplayFormat("yyyy-MM-dd")
        self.to_date.setStyleSheet("""
            QDateEdit {
                background-color: white;
                border: 2px solid #87CEEB;
                border-radius: 5px;
                padding: 6px;
                font-size: 13px;
            }
        """)
        self.to_date.hide()
        tbar.addWidget(self.to_date)
        
        # SEARCH Button (Blue)
        self.search_btn = QPushButton("SEARCH")
        self.search_btn.setStyleSheet("""
            QPushButton {
                background-color: #007BFF;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px 16px;
                border-radius: 6px;
                min-width: 100px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #0056B3;
            }
        """)
        self.search_btn.clicked.connect(self.apply_filters)
        tbar.addWidget(self.search_btn)
        
        # RESET Button (Gray)
        self.reset_btn = QPushButton("RESET")
        self.reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #6C757D;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px 16px;
                border-radius: 6px;
                min-width: 100px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #545B62;
            }
        """)
        self.reset_btn.clicked.connect(self.reset_filters)
        tbar.addWidget(self.reset_btn)
        
        # Download Excel Button (Purple)
        self.dl_btn = QPushButton("Download Excel")
        self.dl_btn.setStyleSheet("""
            QPushButton {
                background-color: #6F42C1;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px 16px;
                border-radius: 6px;
                min-width: 130px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #5A32A3;
            }
        """)
        self.dl_btn.clicked.connect(self.download_excel)
        tbar.addWidget(self.dl_btn)
        
        tbar.addStretch()
        vbox.addLayout(tbar)

        # ------------ table with sky blue headers ------------
        headers = ["Date","Vehicle No","FC Exp","Tyre Amt","Tyre Type",
                   "Tax","Tax Type","Spare Work","Type","Loan","Ins.","Others",
                   "Remarks","Total","Edit","Del"]
        self.table = QTableWidget(0, len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        # Sky blue header styling
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #E0E0E0;
                color: #003366;
                border: 1px solid #87CEEB;
                font-size: 13px;
            }
            QHeaderView::section {
                background-color: #87CEEB;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px;
                border: 1px solid #5DADE2;
            }
            QTableWidget::item:selected {
                background-color: #D4E6F1;
                color: #003366;
            }
        """)
        
        h = self.table.horizontalHeader()
        h.setStretchLastSection(True)
        for i in range(len(headers)):
            h.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        vbox.addWidget(self.table)

        self.total_lbl = QLabel("Grand Total: ₹0")
        self.total_lbl.setStyleSheet("""
            font-weight: bold;
            font-size: 18px;
            color: #003366;
            background-color: #E8F4F8;
            padding: 10px;
            border-radius: 5px;
            border: 2px solid #87CEEB;
        """)
        vbox.addWidget(self.total_lbl)

    def _date_combo_changed(self, txt):
        vis = txt == "Custom"
        self.from_date.setVisible(vis)
        self.to_date.setVisible(vis)
        self.to_label.setVisible(vis)  # Also show/hide the "to" label

    def load_from_db(self):
        """
        Load rows from DB into self.records (list-of-lists) and self.record_ids
        """
        rows = self.db.loadvehicleexpenses()
        self.records = []
        self.record_ids = []
        for r in rows:
            # keep columns consistent with table order in UI:
            rec = [
                r['date'] or "",
                r['vehicle_no'] or "",
                str(r['fc_expense'] or 0),
                str(r['tyre_amount'] or 0),
                r['tyre_type'] or "",
                str(r['tax'] or 0),
                r['tax_type'] or "",
                str(r['spare_work'] or 0),
                r['spare_type'] or "",
                str(r['loan'] or 0),
                str(r['insurance'] or 0),
                str(r['others'] or 0),
                r['remarks'] or ""
            ]
            self.records.append(rec)
            self.record_ids.append(r['id'])
        self.refresh()

    def new_record(self):
        dlg = ExpenseDialog(self)
        if dlg.exec():
            data = dlg.data()  # list of 13 fields (no total)
            if self.db:
                self.db.save_vehicle_expense(data)   # DB computes total
                self.load_from_db()
            else:
                # keep old in-memory behavior
                self.records.insert(0, data)
                self.refresh()

    def edit_record(self, idx):
        # idx is index into self.records
        if idx < 0 or idx >= len(self.records):
            return
        current = self.records[idx]
        dlg = ExpenseDialog(self, current)
        if dlg.exec():
            newdata = dlg.data()
            if self.db:
                expid = self.record_ids[idx]
                self.db.update_vehicle_expense(expid, newdata)
                self.load_from_db()
            else:
                self.records[idx] = newdata
                self.refresh()

    def del_record(self, idx):
        if idx < 0 or idx >= len(self.records):
            return
        # confirm
        if QMessageBox.question(self, "Confirm", "Delete this record?",
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                                ) != QMessageBox.StandardButton.Yes:
            return
        if self.db:
            expid = self.record_ids[idx]
            self.db.delete_vehicle_expense(expid)
            self.load_from_db()
        else:
            del self.records[idx]
            self.refresh()

    # ------------------------------------------------------------------
    def _date_combo_changed(self, txt):
        vis = txt == "Custom"
        self.from_date.setVisible(vis); self.to_date.setVisible(vis)

    # ------------------------------------------------------------------
    def refresh(self, subset=None):
        rows = subset if subset is not None else self.records
        self.table.setRowCount(len(rows))
        grand = 0.0
        for r, rec in enumerate(rows):
            # fill normal columns (0-12)
            for c in range(13):
                itm = QTableWidgetItem(rec[c] if c < len(rec) else "")
                itm.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                itm.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.table.setItem(r, c, itm)

            # compute total
            tot = sum(safe_float(rec[i]) for i in (2,3,5,7,9,10,11))
            grand += tot
            tot_itm = QTableWidgetItem(show_int_amount(tot))
            tot_itm.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            tot_itm.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
            self.table.setItem(r, 13, tot_itm)

            # edit / delete buttons
            e_btn = QPushButton("Edit"); d_btn = QPushButton("X"); d_btn.setStyleSheet("color:red;")
            e_btn.clicked.connect(lambda _=None, i=r: self.edit_record(i))
            d_btn.clicked.connect(lambda _=None, i=r: self.del_record(i))
            self.table.setCellWidget(r, 14, e_btn); self.table.setCellWidget(r, 15, d_btn)

        self.total_lbl.setText(f"Grand Total: ₹{show_int_amount(grand)}")

    # ------------------------------------------------------------------
    # Filtering helpers
    # ------------------------------------------------------------------
    def _date_ok(self, rec_date: QDate, start: QDate | None, end: QDate | None):
        if not start: return True
        return start <= rec_date <= (end or start)

    def apply_filters(self):
        veh = self.filt_vehicle.text().lower().strip()
        choice = self.date_combo.currentText()
        today = QDate.currentDate()
        start = end = None
        if choice == "Last 1 Day":  start = today.addDays(-1);  end = today
        elif choice == "Last Month": start = today.addMonths(-1); end = today
        elif choice == "Last 3 Months": start = today.addMonths(-3); end = today
        elif choice == "Last 6 Months": start = today.addMonths(-6); end = today
        elif choice == "Last Year": start = today.addYears(-1); end = today
        elif choice == "Custom": start, end = self.from_date.date(), self.to_date.date()

        subset = []
        for rec in self.records:
            rec_date = QDate.fromString(rec[0], "yyyy-MM-dd")
            if veh and veh not in rec[1].lower(): continue
            if not self._date_ok(rec_date, start, end): continue
            subset.append(rec)
        self.refresh(subset)

    def reset_filters(self):
        self.filt_vehicle.clear(); self.date_combo.setCurrentIndex(0)
        self.from_date.hide(); self.to_date.hide()
        self.refresh()

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def download_excel(self):
        if not self.records:
            QMessageBox.information(self, "No data", "No records to export.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            # Calculate column totals
            total_fc = sum(safe_float(rec[2]) for rec in self.records if len(rec) > 2)
            total_tyre = sum(safe_float(rec[3]) for rec in self.records if len(rec) > 3)
            total_tax = sum(safe_float(rec[5]) for rec in self.records if len(rec) > 5)
            total_spare = sum(safe_float(rec[7]) for rec in self.records if len(rec) > 7)
            total_loan = sum(safe_float(rec[9]) for rec in self.records if len(rec) > 9)
            total_insurance = sum(safe_float(rec[10]) for rec in self.records if len(rec) > 10)
            total_others = sum(safe_float(rec[11]) for rec in self.records if len(rec) > 11)
            grand_total = total_fc + total_tyre + total_tax + total_spare + total_loan + total_insurance + total_others
            
            # openpyxl first
            if 'openpyxl' in sys.modules:
                wb = Workbook()
                ws = wb.active
                ws.title = "Vehicle Expenses"
                cols = ["Date","Vehicle","FC","Tyre Amt","Tyre Type","Tax","Tax Type",
                        "Spare","Type","Loan","Insurance","Others","Remarks","Total"]
                for c, h in enumerate(cols, 1):
                    cell = ws.cell(1, c, h)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                
                for r, rec in enumerate(self.records, start=2):
                    val = []
                    for i in range(13):
                        val.append(rec[i] if i < len(rec) else "")
                    val.append(sum(safe_float(rec[i]) for i in (2,3,5,7,9,10,11)))
                    for c, v in enumerate(val, 1):
                        ws.cell(r, c, v)
                
                # Add totals row
                total_row = len(self.records) + 2
                ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
                ws.cell(total_row, 3, total_fc)
                ws.cell(total_row, 4, total_tyre)
                ws.cell(total_row, 6, total_tax)
                ws.cell(total_row, 8, total_spare)
                ws.cell(total_row, 10, total_loan)
                ws.cell(total_row, 11, total_insurance)
                ws.cell(total_row, 12, total_others)
                ws.cell(total_row, 14, grand_total).font = Font(bold=True)
                
                wb.save(path)
            else:
                # pandas fallback
                import pandas as pd
                rows = []
                for rec in self.records:
                    val = rec[:13]
                    val += [sum(safe_float(rec[i]) for i in (2,3,5,7,9,10,11))]
                    rows.append(val)
                
                # Add totals row
                totals_row = ["TOTAL", "", total_fc, total_tyre, "", total_tax, "", 
                             total_spare, "", total_loan, total_insurance, total_others, "", grand_total]
                rows.append(totals_row)
                
                df = pd.DataFrame(rows, columns=
                    ["Date","Vehicle","FC","Tyre Amt","Tyre Type","Tax","Tax Type",
                     "Spare","Type","Loan","Insurance","Others","Remarks","Total"])
                df.to_excel(path, index=False)
            
            QMessageBox.information(self, "Saved", f"File saved to\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export:\n{e}")
# ----------------------------------------------------------------------
# VEHICLE / DRIVER  PAGE
# ----------------------------------------------------------------------


class VehicleDialog(QDialog):
    def __init__(self, parent=None, vehicle_name=""):
        super().__init__(parent)
        self.setWindowTitle("Edit Vehicle Details" if vehicle_name else "Add New Vehicle")
        self.setMinimumSize(420, 180)
        self.vehicle_name = vehicle_name
        
        layout = QVBoxLayout(self)
        label = QLabel(f"Editing: {vehicle_name}" if vehicle_name else "Add New Vehicle")
        label.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(label)
        
        self.name_edit = QLineEdit(self)
        self.name_edit.setText(vehicle_name)
        self.name_edit.setPlaceholderText("Enter vehicle name")
        self.name_edit.setMinimumHeight(30)
        layout.addWidget(self.name_edit)
        
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Save", self)
        save_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("Cancel", self)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f4f8;
                color: #003366;
            }
            QPushButton {
                min-width: 80px;
                min-height: 30px;
                background-color: #007bff;
                color: white;
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 5px;
                color: #003366;
                padding: 5px;
                font-size: 16px;
            }
        """)
    
    def get_name(self):
        return self.name_edit.text().strip()

class VehicleWidget(QWidget):
    def __init__(self, name, on_edit, on_remove, on_details, db_manager=None):
        super().__init__()
        self.name = name
        self.on_edit = on_edit
        self.on_remove = on_remove
        self.on_details = on_details
        self.db = db_manager
        self.details_data = {
            'vehicle_no': name,
            'registration_date': '',
            'fitness_upto': '',
            'tax_upto': '',
            'insurance_upto': '',
            'pucc_upto': '',
            'permit_upto': '',
            'national_permit_upto': '',
            'driver_name': '',
            'driver_contact': '',
            'driver_alt_contact': '',
            'driver_experience': '',
            'driver_adhar': '',
            'driver_license_path': '',
            'loan_total': '',
            'loan_paid': '',
            'loan_remaining': '',
            'driver_date_of_joining': '',
            'driver_bank_account': ''
        }
        if self.db:
            self.load_details_from_db()
        self.init_ui()
    
    def load_details_from_db(self):
        if self.db:
            data = self.db.load_vehicle_driver_details(self.name)
            if data:
                self.details_data.update(data)
    
    def init_ui(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #e9f0fb;
                border-radius: 10px;
                padding: 12px;
            }
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                font-size: 14px;
                padding: 6px 12px;
                border-radius: 6px;
                min-width: 70px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QLabel#titleLabel {
                color: #003366;
                font-size: 16px;
                font-weight: bold;
            }
            QLabel#summaryLabel {
                color: #30475e;
                font-size: 11px;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        
        self.label = QLabel(self.name, self)
        self.label.setObjectName("titleLabel")
        layout.addWidget(self.label)
        
        self.details_summary = QLabel("(no details)", self)
        self.details_summary.setObjectName("summaryLabel")
        layout.addWidget(self.details_summary)
        
        btn_layout = QHBoxLayout()
        
        details_btn = QPushButton("Details", self)
        details_btn.clicked.connect(self.details_clicked)
        btn_layout.addWidget(details_btn)
        
        edit_btn = QPushButton("Edit", self)
        edit_btn.clicked.connect(self.edit_clicked)
        btn_layout.addWidget(edit_btn)
        
        remove_btn = QPushButton("Remove", self)
        remove_btn.clicked.connect(self.remove_clicked)
        btn_layout.addWidget(remove_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.setMinimumWidth(300)
        self.setMaximumWidth(380)
        self.update_details_summary() # Call here to show initial summary

    def edit_clicked(self):
        self.on_edit(self)
    
    def remove_clicked(self):
        self.on_remove(self)
    
    def details_clicked(self):
        self.on_details(self)
    
    def update_name(self, new_name):
        # If vehicle_no is a primary key or unique, handle update carefully
        # For now, assume it's just updating the display name and internal data
        old_name = self.name
        self.name = new_name
        self.label.setText(new_name)
        self.details_data['vehicle_no'] = new_name
        if self.db:
            # This is a simplified update. In a real scenario, you might need to
            # delete the old record and insert a new one if vehicle_no is a primary key,
            # or update the vehicle_no field if it's just a unique identifier.
            # For this example, we'll assume the DB update handles the unique constraint.
            # A more robust solution would be to pass the old_name to the DB manager
            # to find and update the record.
            existing_data = self.db.load_vehicle_driver_details(old_name)
            if existing_data:
                self.db.delete_vehicle_driver_details(old_name) # Delete old record
                self.db.save_vehicle_driver_details(self.details_data) # Save new record with updated name
            else:
                self.db.save_vehicle_driver_details(self.details_data) # Save new record if not found
        self.update_details_summary()
    
    def update_details_summary(self):
        d = self.details_data
        parts = []
        if d.get('registration_date'):
            parts.append(f"Reg: {d['registration_date']}")
        if d.get('fitness_upto'):
            parts.append(f"Fitness: {d['fitness_upto']}")
        if d.get('tax_upto'):
            parts.append(f"Tax: {d['tax_upto']}")
        
        if parts:
            self.details_summary.setText(" | ".join(parts))
        else:
            self.details_summary.setText("(no details)")

class VehicleDriverPage(QWidget):
    def __init__(self, back_callback=None, db_manager=None):
        super().__init__()
        self.back_callback = back_callback
        self.db = db_manager
        self.setWindowTitle("Driver & Vehicle Management")
        self.setMinimumSize(900, 640)
        self.vehicles = []
        self.init_ui()
        if self.db:
            self.load_vehicles_from_db()
    
    def init_ui(self):
        main_layout = QVBoxLayout(self)
        
        # Header with back button
        header_layout = QHBoxLayout()
        
        # Back button
        back_btn = QPushButton("Back to Home", self)
        back_btn.setStyleSheet("""
            QPushButton {
                background-color: #336699;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border-radius: 8px;
                font-size: 14px;
                border: none;
            }
            QPushButton:hover {
                background-color: #294d66;
            }
        """)
        back_btn.clicked.connect(self.go_back_home)
        header_layout.addWidget(back_btn)
        
        header_layout.addStretch()
        
        # Page title
        title_label = QLabel("Driver & Vehicle Management")
        title_label.setFont(QFont("Arial Black", 24))
        title_label.setStyleSheet("color: #003366; font-weight: bold;")
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        add_btn = QPushButton("+ New Vehicle", self)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                font-weight: bold;
                padding: 8px 14px;
                border-radius: 6px;
                border: none;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        add_btn.clicked.connect(self.add_vehicle)
        header_layout.addWidget(add_btn)
        
        main_layout.addLayout(header_layout)
        
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        main_layout.addWidget(self.scroll_area)
        
        self.container = QWidget()
        self.grid_layout = QGridLayout(self.container)
        self.grid_layout.setSpacing(18)
        self.grid_layout.setContentsMargins(12, 12, 12, 12)
        self.scroll_area.setWidget(self.container)
        
    def load_vehicles_from_db(self):
        if self.db:
            all_details = self.db.load_all_vehicle_driver_details()
            self.vehicles = [] # Clear existing in-memory list
            for detail in all_details:
                vehicle_widget = VehicleWidget(detail['vehicle_no'], self.edit_vehicle, self.remove_vehicle, self.show_vehicle_details, self.db)
                vehicle_widget.details_data.update(detail) # Load all details
                self.vehicles.append(vehicle_widget)
            self.refresh_grid()

    def go_back_home(self):
        if self.back_callback:
            self.back_callback()
    
    def create_vehicle(self, name):
        vehicle_widget = VehicleWidget(name, self.edit_vehicle, self.remove_vehicle, self.show_vehicle_details, self.db)
        self.vehicles.append(vehicle_widget)
        if self.db:
            # Save initial vehicle_no to DB
            initial_data = {'vehicle_no': name}
            # Fill other fields with defaults if not present in initial_data
            for key in vehicle_widget.details_data:
                if key not in initial_data:
                    initial_data[key] = vehicle_widget.details_data[key]
            self.db.save_vehicle_driver_details(initial_data)
            vehicle_widget.load_details_from_db() # Reload to get any default values/ID
        vehicle_widget.update_details_summary()
        self.refresh_grid()
    
    def refresh_grid(self):
        for i in reversed(range(self.grid_layout.count())):
            widget = self.grid_layout.itemAt(i).widget()
            if widget:
                self.grid_layout.removeWidget(widget)
                widget.setParent(None)
        
        for idx, vehicle in enumerate(self.vehicles):
            row = idx // 2
            col = idx % 2
            self.grid_layout.addWidget(vehicle, row, col)
        
        self.grid_layout.setRowStretch((len(self.vehicles) + 1) // 2, 1)
    
    def add_vehicle(self):
        dlg = VehicleDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            name = dlg.get_name()
            if name:
                # Check if vehicle already exists
                if any(v.name == name for v in self.vehicles):
                    QMessageBox.warning(self, "Warning", f"Vehicle '{name}' already exists.")
                    return
                self.create_vehicle(name)
            else:
                QMessageBox.warning(self, "Warning", "Vehicle name cannot be empty.")
    
    def edit_vehicle(self, vehicle_widget):
        dlg = VehicleDialog(self, vehicle_widget.name)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_name = dlg.get_name()
            if new_name:
                if new_name != vehicle_widget.name and any(v.name == new_name for v in self.vehicles):
                    QMessageBox.warning(self, "Warning", f"Vehicle '{new_name}' already exists.")
                    return
                vehicle_widget.update_name(new_name)
            else:
                QMessageBox.warning(self, "Warning", "Vehicle name cannot be empty.")
    
    def remove_vehicle(self, vehicle_widget):
        confirm = QMessageBox.question(
            self, "Confirm Remove", f"Remove {vehicle_widget.name}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if confirm == QMessageBox.StandardButton.Yes:
            if self.db:
                self.db.delete_vehicle_driver_details(vehicle_widget.name)
            self.vehicles.remove(vehicle_widget)
            self.refresh_grid()
    
    def show_vehicle_details(self, vehicle_widget):
        """FIXED: Comprehensive vehicle details dialog with proper scrollable layout"""
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Vehicle Details - {vehicle_widget.name}")
        dlg.setMinimumSize(750, 600)  # Made wider for better layout
        dlg.resize(750, 600)  # Set initial size
        
        # Main layout for the dialog
        main_layout = QVBoxLayout(dlg)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(10)
        
        # Title
        title_label = QLabel(f"Vehicle Details - {vehicle_widget.name}")
        title_label.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title_label.setStyleSheet("color: #003366; padding-bottom: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title_label)
        
        # FIXED: Create scroll area for the form content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # Content widget inside scroll area
        scroll_content = QWidget()
        content_layout = QVBoxLayout(scroll_content)
        content_layout.setContentsMargins(10, 10, 10, 10)
        content_layout.setSpacing(15)
        
        # Helper function for date edit
        def make_dateedit_from_str(date_str):
            d = QDate.currentDate()
            if date_str:
                qd = QDate.fromString(date_str, "yyyy-MM-dd")
                if qd.isValid():
                    d = qd
            de = QDateEdit()
            de.setCalendarPopup(True)
            de.setDisplayFormat("yyyy-MM-dd")
            de.setDate(d)
            de.setMinimumHeight(30)
            return de
        
        # SECTION 1: Vehicle Information
        vehicle_group = QGroupBox("Vehicle Information")
        vehicle_group.setStyleSheet("QGroupBox { font-weight: bold; color: #003366; border: 2px solid #336699; border-radius: 8px; margin: 10px 0; padding-top: 15px; }")
        vehicle_layout = QFormLayout(vehicle_group)
        vehicle_layout.setSpacing(10)
        
        vehicle_no_edit = QLineEdit()
        vehicle_no_edit.setText(vehicle_widget.details_data.get('vehicle_no', vehicle_widget.name))
        vehicle_no_edit.setMinimumHeight(30)
        vehicle_layout.addRow("Vehicle No:", vehicle_no_edit)
        
        reg_de = make_dateedit_from_str(vehicle_widget.details_data.get('registration_date', ''))
        vehicle_layout.addRow("Registration Date:", reg_de)
        
        fitness_de = make_dateedit_from_str(vehicle_widget.details_data.get('fitness_upto', ''))
        vehicle_layout.addRow("Fitness Valid UpTo:", fitness_de)
        
        tax_de = make_dateedit_from_str(vehicle_widget.details_data.get('tax_upto', ''))
        vehicle_layout.addRow("Tax Valid UpTo:", tax_de)
        
        insurance_de = make_dateedit_from_str(vehicle_widget.details_data.get('insurance_upto', ''))
        vehicle_layout.addRow("Insurance Valid UpTo:", insurance_de)
        
        pucc_de = make_dateedit_from_str(vehicle_widget.details_data.get('pucc_upto', ''))
        vehicle_layout.addRow("PUCC Valid UpTo:", pucc_de)
        
        permit_de = make_dateedit_from_str(vehicle_widget.details_data.get('permit_upto', ''))
        vehicle_layout.addRow("Permit Valid UpTo:", permit_de)
        
        nat_permit_de = make_dateedit_from_str(vehicle_widget.details_data.get('national_permit_upto', ''))
        vehicle_layout.addRow("National Permit Valid UpTo:", nat_permit_de)
        
        content_layout.addWidget(vehicle_group)
        
        # SECTION 2: Driver Information
        driver_group = QGroupBox("Driver Information")
        driver_group.setStyleSheet("QGroupBox { font-weight: bold; color: #003366; border: 2px solid #336699; border-radius: 8px; margin: 10px 0; padding-top: 15px; }")
        driver_layout = QFormLayout(driver_group)
        driver_layout.setSpacing(10)
        
        driver_name_edit = QLineEdit()
        driver_name_edit.setText(vehicle_widget.details_data.get('driver_name', ''))
        driver_name_edit.setMinimumHeight(30)
        driver_layout.addRow("Driver Name:", driver_name_edit)
        
        driver_contact_edit = QLineEdit()
        driver_contact_edit.setText(vehicle_widget.details_data.get('driver_contact', ''))
        driver_contact_edit.setMinimumHeight(30)
        driver_layout.addRow("Contact No:", driver_contact_edit)
        
        driver_alt_contact_edit = QLineEdit()
        driver_alt_contact_edit.setText(vehicle_widget.details_data.get('driver_alt_contact', ''))
        driver_alt_contact_edit.setMinimumHeight(30)
        driver_layout.addRow("Alternative Number:", driver_alt_contact_edit)
        
        driver_experience_edit = QLineEdit()
        driver_experience_edit.setText(vehicle_widget.details_data.get('driver_experience', ''))
        driver_experience_edit.setMinimumHeight(30)
        driver_layout.addRow("Experience (Years):", driver_experience_edit)
        
        driver_adhar_edit = QLineEdit()
        driver_adhar_edit.setText(vehicle_widget.details_data.get('driver_adhar', ''))
        driver_adhar_edit.setMinimumHeight(30)
        driver_layout.addRow("Aadhaar Number:", driver_adhar_edit)
        
        date_of_joining_de = make_dateedit_from_str(vehicle_widget.details_data.get('driver_date_of_joining', ''))
        driver_layout.addRow("Date of Joining:", date_of_joining_de)
        
        bank_account_edit = QLineEdit()
        bank_account_edit.setText(vehicle_widget.details_data.get('driver_bank_account', ''))
        bank_account_edit.setMinimumHeight(30)
        driver_layout.addRow("Bank Account Number:", bank_account_edit)
        
        content_layout.addWidget(driver_group)
        
        # SECTION 3: FIXED License Upload Section
        license_group = QGroupBox("Driver License")
        license_group.setStyleSheet("QGroupBox { font-weight: bold; color: #003366; border: 2px solid #336699; border-radius: 8px; margin: 10px 0; padding-top: 15px; }")
        license_main_layout = QVBoxLayout(license_group)
        license_main_layout.setSpacing(10)
        
        # License upload controls
        license_controls_layout = QHBoxLayout()
        
        license_upload_btn = QPushButton("Upload License")
        license_upload_btn.setMinimumSize(120, 35)
        license_upload_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        
        license_remove_btn = QPushButton("Remove License")
        license_remove_btn.setMinimumSize(120, 35)
        license_remove_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                font-weight: bold;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        license_remove_btn.setEnabled(bool(vehicle_widget.details_data.get('driver_license_path', '')))
        
        license_path_label = QLabel()
        current_path = vehicle_widget.details_data.get('driver_license_path', '')
        if current_path:
            license_path_label.setText(f"File: {os.path.basename(current_path)}")
        else:
            license_path_label.setText("No file uploaded")
        license_path_label.setStyleSheet("font-style: italic; color: #555555; padding: 5px;")
        license_path_label.setWordWrap(True)
        
        license_controls_layout.addWidget(license_upload_btn)
        license_controls_layout.addWidget(license_remove_btn)
        license_controls_layout.addWidget(license_path_label, 1)  # Give it more space
        license_main_layout.addLayout(license_controls_layout)
        
        # FIXED: License image display area
        license_image_container = QFrame()
        license_image_container.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Sunken)
        license_image_container.setStyleSheet("border: 2px solid #ccc; background-color: white; border-radius: 8px;")
        license_image_container.setMinimumHeight(220)  # Fixed height
        license_image_container.setMaximumHeight(220)  # Prevent expansion
        
        license_image_layout = QVBoxLayout(license_image_container)
        license_image_layout.setContentsMargins(5, 5, 5, 5)
        
        license_image_label = QLabel()
        license_image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        license_image_label.setMinimumSize(300, 200)
        license_image_label.setMaximumSize(300, 200)  # Fixed size
        license_image_label.setScaledContents(False)  # Prevent distortion
        license_image_label.setStyleSheet("border: 1px solid #ddd; background-color: #f9f9f9;")
        
        license_image_layout.addWidget(license_image_label, alignment=Qt.AlignmentFlag.AlignCenter)
        license_main_layout.addWidget(license_image_container)
        
        content_layout.addWidget(license_group)
        
        # Image display function
        def show_license_image(path):
            """Safely display license image with proper sizing"""
            try:
                if not path or not os.path.isfile(path):
                    license_image_label.clear()
                    license_image_label.setText("No Image Available")
                    license_image_label.setStyleSheet("border: 1px solid #ddd; background-color: #f9f9f9; color: #666;")
                    return
                
                # Verify file is a supported image format
                supported_formats = []
                try:
                    for fmt in QImageReader.supportedImageFormats():
                        supported_formats.append(bytes(fmt).decode().lower())
                except Exception as e:
                    print(f"Error getting supported formats: {e}")
                    supported_formats = ['jpg', 'jpeg', 'png', 'bmp', 'gif']
                
                file_ext = os.path.splitext(path)[1][1:].lower()
                if file_ext not in supported_formats:
                    license_image_label.clear()
                    license_image_label.setText("Unsupported Image Format")
                    license_image_label.setStyleSheet("border: 1px solid #ddd; background-color: #fff3cd; color: #856404;")
                    return
                
                # Load and display the image
                pixmap = QPixmap(path)
                if pixmap.isNull():
                    license_image_label.clear()
                    license_image_label.setText("Invalid Image File")
                    license_image_label.setStyleSheet("border: 1px solid #ddd; background-color: #f8d7da; color: #721c24;")
                    return
                
                # Scale image to fit properly
                scaled_pixmap = pixmap.scaled(
                    license_image_label.width() - 10, license_image_label.height() - 10,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation
                )
                license_image_label.setPixmap(scaled_pixmap)
                license_image_label.setStyleSheet("border: 1px solid #ddd; background-color: white;")
                
            except Exception as e:
                print(f"Error loading image: {e}")
                license_image_label.clear()
                license_image_label.setText("Error Loading Image")
                license_image_label.setStyleSheet("border: 1px solid #ddd; background-color: #f8d7da; color: #721c24;")
        
        # Display existing image
        show_license_image(vehicle_widget.details_data.get('driver_license_path', ''))
        
        # Upload function
        def upload_license():
            try:
                fname, _ = QFileDialog.getOpenFileName(
                    dlg, 
                    "Select License Image", 
                    "", 
                    "Image Files (*.png *.jpg *.jpeg *.bmp *.gif *.tiff);;All Files (*)"
                )
                
                if not fname:
                    return
                
                # Validate file
                if not os.path.isfile(fname):
                    QMessageBox.warning(dlg, "Error", "Selected file does not exist.")
                    return
                
                # Check file size (limit to 10MB)
                try:
                    file_size = os.path.getsize(fname)
                    if file_size > 10 * 1024 * 1024:
                        QMessageBox.warning(dlg, "Error", "File size too large. Please select a file smaller than 10MB.")
                        return
                except Exception as e:
                    print(f"Error checking file size: {e}")
                
                # Verify it's a valid image
                test_pixmap = QPixmap(fname)
                if test_pixmap.isNull():
                    QMessageBox.warning(dlg, "Error", "Selected file is not a valid image.")
                    return
                
                # Save and update UI
                vehicle_widget.details_data['driver_license_path'] = fname
                license_path_label.setText(f"File: {os.path.basename(fname)}")
                license_remove_btn.setEnabled(True)
                show_license_image(fname)
                
                QMessageBox.information(dlg, "Success", "License image uploaded successfully!")
                
            except Exception as e:
                print(f"Error uploading image: {e}")
                QMessageBox.critical(dlg, "Upload Error", f"Failed to upload image: {str(e)}")
        
        # Remove function
        def remove_license():
            try:
                confirm = QMessageBox.question(
                    dlg,
                    "Confirm Remove License",
                    "Are you sure you want to remove the uploaded license image?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if confirm == QMessageBox.StandardButton.Yes:
                    vehicle_widget.details_data['driver_license_path'] = ''
                    license_path_label.setText('No file uploaded')
                    license_remove_btn.setEnabled(False)
                    show_license_image('')
                    QMessageBox.information(dlg, "Removed", "License image removed successfully!")
                    
            except Exception as e:
                print(f"Error removing image: {e}")
                QMessageBox.critical(dlg, "Remove Error", f"Failed to remove image: {str(e)}")
        
        license_upload_btn.clicked.connect(upload_license)
        license_remove_btn.clicked.connect(remove_license)
        
        # SECTION 4: FIXED Loan Information Section
        loan_group = QGroupBox("Loan Information")
        loan_group.setStyleSheet("QGroupBox { font-weight: bold; color: #003366; border: 2px solid #336699; border-radius: 8px; margin: 10px 0; padding-top: 15px; }")
        loan_layout = QVBoxLayout(loan_group)
        loan_layout.setSpacing(15)
        
        # Loan input row
        loan_input_layout = QHBoxLayout()
        
        loan_total_edit = QLineEdit()
        loan_total_edit.setText(str(vehicle_widget.details_data.get('loan_total', '')))
        loan_total_edit.setPlaceholderText("Enter total loan amount")
        loan_total_edit.setMinimumHeight(35)
        
        loan_paid_edit = QLineEdit()
        loan_paid_edit.setText(str(vehicle_widget.details_data.get('loan_paid', '')))
        loan_paid_edit.setPlaceholderText("Enter paid amount")
        loan_paid_edit.setMinimumHeight(35)
        
        loan_input_layout.addWidget(QLabel("Total Loan Amount:"))
        loan_input_layout.addWidget(loan_total_edit)
        loan_input_layout.addWidget(QLabel("Paid Amount:"))
        loan_input_layout.addWidget(loan_paid_edit)
        
        loan_layout.addLayout(loan_input_layout)
        
        # Remaining amount display
        remaining_layout = QHBoxLayout()
        loan_remaining_label = QLabel(str(vehicle_widget.details_data.get('loan_remaining', '0.00')))
        loan_remaining_label.setStyleSheet("""
            QLabel {
                background-color: #e7f3ff;
                border: 2px solid #2196F3;
                border-radius: 8px;
                padding: 10px;
                font-size: 16px;
                font-weight: bold;
                color: #1976D2;
            }
        """)
        loan_remaining_label.setMinimumHeight(40)
        loan_remaining_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        remaining_layout.addWidget(QLabel("Remaining Amount:"))
        remaining_layout.addWidget(loan_remaining_label, 1)
        
        loan_layout.addLayout(remaining_layout)
        content_layout.addWidget(loan_group)
        
        # Loan calculation function
        def update_remaining():
            try:
                total = safe_float(loan_total_edit.text())
                paid = safe_float(loan_paid_edit.text())
                remaining = max(total - paid, 0)
                loan_remaining_label.setText(f"₹ {remaining:,.2f}")
            except Exception as e:
                print(f"Error calculating remaining: {e}")
                loan_remaining_label.setText("₹ 0.00")
        
        loan_total_edit.textChanged.connect(update_remaining)
        loan_paid_edit.textChanged.connect(update_remaining)
        update_remaining()
        
        # Set the content widget in scroll area
        scroll_area.setWidget(scroll_content)
        main_layout.addWidget(scroll_area)
        
        # FIXED: Bottom buttons outside scroll area
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        save_btn = QPushButton("Save Changes")
        save_btn.setMinimumSize(120, 40)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setMinimumSize(120, 40)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        main_layout.addLayout(button_layout)
        
        # Save function
        def on_save():
            try:
                # Save all form data
                vehicle_widget.details_data['vehicle_no'] = vehicle_no_edit.text().strip()
                vehicle_widget.details_data['registration_date'] = reg_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['fitness_upto'] = fitness_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['tax_upto'] = tax_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['insurance_upto'] = insurance_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['pucc_upto'] = pucc_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['permit_upto'] = permit_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['national_permit_upto'] = nat_permit_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['driver_name'] = driver_name_edit.text().strip()
                vehicle_widget.details_data['driver_contact'] = driver_contact_edit.text().strip()
                vehicle_widget.details_data['driver_alt_contact'] = driver_alt_contact_edit.text().strip()
                vehicle_widget.details_data['driver_experience'] = driver_experience_edit.text().strip()
                vehicle_widget.details_data['driver_adhar'] = driver_adhar_edit.text().strip()
                vehicle_widget.details_data['driver_date_of_joining'] = date_of_joining_de.date().toString("yyyy-MM-dd")
                vehicle_widget.details_data['driver_bank_account'] = bank_account_edit.text().strip()
                vehicle_widget.details_data['loan_total'] = safe_float(loan_total_edit.text())
                vehicle_widget.details_data['loan_paid'] = safe_float(loan_paid_edit.text())
                vehicle_widget.details_data['loan_remaining'] = safe_float(loan_remaining_label.text().replace('₹', '').replace(',', '').strip())
                
                if self.db:
                    self.db.save_vehicle_driver_details(vehicle_widget.details_data) # This handles both insert and update
                
                vehicle_widget.update_details_summary()
                QMessageBox.information(dlg, "Success", f"Details saved successfully for {vehicle_widget.name}")
                dlg.accept()
                
            except Exception as e:
                print(f"Error saving: {e}")
                QMessageBox.critical(dlg, "Save Error", f"Failed to save: {str(e)}")
        
        save_btn.clicked.connect(on_save)
        cancel_btn.clicked.connect(dlg.reject)
        
        # Apply comprehensive styling to dialog
        dlg.setStyleSheet("""
            QDialog { 
                background-color: #f8f9fa; 
                color: #003366; 
            }
            QGroupBox { 
                font-size: 14px; 
                font-weight: bold; 
                color: #003366; 
                border: 2px solid #336699; 
                border-radius: 10px; 
                margin-top: 15px; 
                padding-top: 15px; 
                background-color: white;
            }
            QGroupBox::title { 
                subcontrol-origin: margin; 
                left: 10px; 
                padding: 0 10px 0 10px; 
                background-color: #f8f9fa;
            }
            QLabel { 
                color: #003366; 
                font-weight: bold; 
                font-size: 13px; 
                padding: 2px; 
            }
            QLineEdit { 
                background-color: white; 
                color: #003366; 
                border: 2px solid #ced4da; 
                border-radius: 6px; 
                padding: 8px; 
                font-size: 13px; 
                min-height: 20px;
            }
            QLineEdit:focus {
                border-color: #336699;
            }
            QDateEdit { 
                background-color: white; 
                color: #003366; 
                border: 2px solid #ced4da; 
                border-radius: 6px; 
                padding: 8px; 
                font-size: 13px; 
                min-height: 20px;
            }
            QDateEdit:focus {
                border-color: #336699;
            }
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 8px;
                background-color: white;
            }
        """)
        
        dlg.exec()
        self.refresh_grid()


# ----------------------------------------------------------------------

# OFFICE EXPENSE PAGE  

# ----------------------------------------------------------------------


class OfficeExpensePage(QWidget):
    def __init__(self, back_callback=None, db_manager=None):
        super().__init__()
        self.back_callback = back_callback
        self.db = db_manager
        self.month_widgets = []
        self.init_ui()
        # Load existing records from database
        if self.db:
            self.load_from_db()

    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("Office Expense Page")
        self.resize(1000, 700)
        layout = QGridLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        self.setStyleSheet("background-color: white;")

        # Initialize summary_labels and business_summary_labels as empty dicts first
        self.summary_labels = {}
        self.business_summary_labels = {}

        back_btn = QPushButton("← Back to Home")
        back_btn.setStyleSheet("""
            QPushButton {
                background: #336699; color: white;
                font-weight: bold; padding: 8px 16px;
                border-radius: 8px;
                font-size: 14px;
            }
            QPushButton:hover { background: #294d66; }
        """)
        back_btn.clicked.connect(self.go_back_home)
        layout.addWidget(back_btn, 0, 0, 1, 1, alignment=Qt.AlignmentFlag.AlignLeft)

        header = QLabel("Office Expense")
        header.setFont(QFont("Arial Black", 30))
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("background-color:#336699; color:white; padding:15px; border-radius:15px;")
        layout.addWidget(header, 0, 1, 1, 1)

        left_vbox = QVBoxLayout()
        label_monthly = QLabel("Monthly Records")
        label_monthly.setFont(QFont("Arial", 18))
        label_monthly.setStyleSheet("color:#003366; padding-bottom:10px;")
        left_vbox.addWidget(label_monthly)

        self.btn_add_month = QPushButton("New Month Record")
        self.btn_add_month.setStyleSheet("""
            QPushButton {
                background: #336699; color: white;
                font-weight: bold; padding: 10px; border-radius: 10px;
                font-size: 16px;
            }
            QPushButton:hover { background: #294d66; }
        """)
        self.btn_add_month.clicked.connect(self.prompt_add_month_record)
        left_vbox.addWidget(self.btn_add_month)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setContentsMargins(5, 5, 5, 5)
        self.scroll_layout.setSpacing(15)
        self.scroll_area.setWidget(self.scroll_content)
        left_vbox.addWidget(self.scroll_area)

        left_widget = QWidget()
        left_widget.setLayout(left_vbox)

        right_vbox = QVBoxLayout()
        label_summary = QLabel("Office Expenses Summary")
        label_summary.setFont(QFont("Arial", 20))
        label_summary.setStyleSheet("color:#003366; margin-bottom: 10px;")
        right_vbox.addWidget(label_summary)

        # Clear and reinitialize to prevent errors
        self.summary_labels.clear()
        summary_items = [
            "Total Current Bill",
            "Manager Salary",
            "Office Rent",
            "Others",
            "Grand Total"
        ]
        for item in summary_items:
            h_layout = QHBoxLayout()
            lbl = QLabel(item)
            lbl.setFont(QFont("Arial", 14))
            lbl.setStyleSheet("color:#222222;")
            val = QLabel("₹ 0.00")
            val.setFont(QFont("Arial", 14))
            val.setStyleSheet("background:#d0e4f7; border:1px solid #336699; border-radius:6px; padding:4px; color:#003366;")
            h_layout.addWidget(lbl)
            h_layout.addWidget(val)
            right_vbox.addLayout(h_layout)
            self.summary_labels[item] = val

        label_business_summary = QLabel("Business Summary")
        label_business_summary.setFont(QFont("Arial", 20))
        label_business_summary.setStyleSheet("color:#003366; margin: 20px 0 10px 0;")
        right_vbox.addWidget(label_business_summary)

        # Clear and reinitialize to prevent errors
        self.business_summary_labels.clear()
        business_items = [
            "Trip Expenses",
            "Trip Profits",
            "Vehicle Expenses",
            "Office Expenses",
            "Grand Total"
        ]
        for item in business_items:
            h_layout = QHBoxLayout()
            lbl = QLabel(item)
            lbl.setFont(QFont("Arial", 14))
            lbl.setStyleSheet("color:#222222;")
            val = QLabel("₹ 0.00")
            val.setFont(QFont("Arial", 14))
            
            # Special styling for Grand Total
            if item == "Grand Total":
                val.setStyleSheet("background:#fff3cd; border:2px solid #856404; border-radius:6px; padding:6px; color:#856404; font-weight:bold; font-size:16px;")
            else:
                val.setStyleSheet("background:#f4e6e6; border:1px solid #993333; border-radius:6px; padding:4px; color:#660000;")
            
            h_layout.addWidget(lbl)
            h_layout.addWidget(val)
            right_vbox.addLayout(h_layout)
            self.business_summary_labels[item] = val

        refresh_btn = QPushButton("Refresh Totals")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background:#336699; color:white;
                padding: 10px; border-radius: 8px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover { background:#294d66; }
        """)
        refresh_btn.clicked.connect(self.refresh_totals)
        right_vbox.addWidget(refresh_btn)
        right_vbox.addStretch(10)

        right_widget = QWidget()
        right_widget.setLayout(right_vbox)

        layout.addWidget(left_widget, 1, 0)
        layout.addWidget(right_widget, 1, 1)
        layout.setColumnStretch(0, 1)
        layout.setColumnStretch(1, 1)

    def load_from_db(self):
        """Load office expense records from database"""
        if not self.db:
            return
        
        try:
            records = self.db.load_office_expenses()
            for record in records:
                month_widget = MonthWidget(
                    record['month'] or f"Month {len(self.month_widgets) + 1}",
                    self.safe_refresh_totals,  # Use safe wrapper
                    parent=self
                )
                # Store database ID
                month_widget.db_id = record['id']
                
                # Load field values
                month_widget.fields["Current Bill"].setText(str(record['current_bill'] or 0))
                month_widget.fields["Manager Salary"].setText(str(record['manager_salary'] or 0))
                month_widget.fields["Office Expenses"].setText(str(record['office_rent'] or 0))
                month_widget.fields["Other Expenses"].setText(str(record['others'] or 0))
                
                month_widget.collapse()
                month_widget.disable_editing()
                self.scroll_layout.insertWidget(0, month_widget)
                self.month_widgets.insert(0, month_widget)
            
            self.refresh_totals()
        except Exception as e:
            print(f"Error in load_from_db: {e}")
            QMessageBox.warning(self, "Database Error", f"Failed to load office expenses: {str(e)}")

    def safe_refresh_totals(self):
        """Safe wrapper for refresh_totals to prevent errors during initialization"""
        try:
            self.refresh_totals()
        except Exception as e:
            print(f"Error in safe_refresh_totals: {e}")

    def prompt_add_month_record(self):
        text, ok = QInputDialog.getText(self, "New Month Record", "Enter name for new month record:")
        if ok and text.strip():
            self.add_month_record(text.strip())

    def add_month_record(self, month_name=None):
        if month_name is None:
            month_name = f"Month {len(self.month_widgets) + 1} - 2025"
        
        # Create new record in database
        if self.db:
            try:
                # Save to database with zero values initially
                data = (month_name, 0.0, 0.0, 0.0, 0.0)
                new_id = self.db.save_office_expense(data)
                
                # Create widget
                month_widget = MonthWidget(month_name, self.safe_refresh_totals, parent=self)  # Use safe wrapper
                month_widget.db_id = new_id
                month_widget.collapse()
                self.scroll_layout.insertWidget(0, month_widget)
                self.month_widgets.insert(0, month_widget)
                self.refresh_totals()
            except Exception as e:
                print(f"Error in add_month_record: {e}")
                QMessageBox.warning(self, "Database Error", f"Failed to add month record: {str(e)}")
        else:
            # Fallback without database
            month_widget = MonthWidget(month_name, self.safe_refresh_totals, parent=self)  # Use safe wrapper
            month_widget.collapse()
            self.scroll_layout.insertWidget(0, month_widget)
            self.month_widgets.insert(0, month_widget)
            self.refresh_totals()

    def save_month_to_db(self, month_widget):
        """Save a month widget's data to database"""
        if not self.db or not hasattr(month_widget, 'db_id'):
            return
        
        try:
            month_name = month_widget.month_edit.text()
            totals = month_widget.get_totals()
            
            if not totals:  # Safety check
                print("Warning: get_totals returned None")
                return
            
            data = (
                month_name,
                totals.get("Current Bill", 0.0),
                totals.get("Manager Salary", 0.0),
                totals.get("Office Expenses", 0.0),
                totals.get("Other Expenses", 0.0)
            )
            
            self.db.update_office_expense(month_widget.db_id, data)
        except Exception as e:
            print(f"Error in save_month_to_db: {e}")
            QMessageBox.warning(self, "Database Error", f"Failed to save record: {str(e)}")

    def delete_month_from_db(self, month_widget):
        """Delete a month widget's data from database"""
        if not self.db or not hasattr(month_widget, 'db_id'):
            return
        
        try:
            self.db.delete_office_expense(month_widget.db_id)
        except Exception as e:
            QMessageBox.warning(self, "Database Error", f"Failed to delete record: {str(e)}")

    def get_trip_totals_from_db(self):
            """Fetch trip expenses and profits from database"""
            if not self.db:
                return 0.0, 0.0
            
            try:
                trips = self.db.load_trips()
                total_expense = 0.0
                total_profit = 0.0
                
                for trip in trips:
                    total_expense += safe_float(trip['expense'] or 0)
                    total_profit += safe_float(trip['profit'] or 0)
                
                return total_expense, total_profit
            except Exception as e:
                print(f"Error fetching trip totals: {e}")
                return 0.0, 0.0

    def get_vehicle_expense_from_db(self):
            """Fetch total vehicle expenses from database"""
            if not self.db:
                return 0.0
            
            try:
                vehicle_expenses = self.db.load_vehicle_expenses()
                total = 0.0
                
                for expense in vehicle_expenses:
                    total += safe_float(expense['total'] or 0)
                
                return total
            except Exception as e:
                print(f"Error fetching vehicle expenses: {e}")
                return 0.0

    def get_office_expense_total(self):
            """Calculate total office expenses from current month widgets"""
            total = 0.0
            
            for widget in self.month_widgets:
                if widget.isVisible():
                    totals = widget.get_totals()
                    if totals:
                        total += (
                            totals.get("Office Expenses", 0.0)
                            + totals.get("Manager Salary", 0.0)
                            + totals.get("Current Bill", 0.0)
                            + totals.get("Other Expenses", 0.0)
                        )
            
            return total
        
    def refresh_totals(self):
            """Refresh all totals - with error handling"""
            try:
                # Clean up widgets that were deleted
                self.month_widgets = [w for w in self.month_widgets if w.parent() is not None]

                total_current_bill = 0.0
                total_manager_salary = 0.0
                total_office_rent = 0.0
                total_others = 0.0
                total_office_expenses = 0.0

                for widget in self.month_widgets:
                    if widget.isVisible():
                        totals = widget.get_totals()
                        if totals:  # Make sure totals is not None
                            total_current_bill += totals.get("Current Bill", 0.0)
                            total_manager_salary += totals.get("Manager Salary", 0.0)
                            total_office_rent += totals.get("Office Expenses", 0.0)
                            total_others += totals.get("Other Expenses", 0.0)

                            total_office_expenses += (
                                totals.get("Office Expenses", 0.0)
                                + totals.get("Manager Salary", 0.0)
                                + totals.get("Current Bill", 0.0)
                                + totals.get("Other Expenses", 0.0)
                            )

                # Update Office Expenses Summary
                grand_total = total_current_bill + total_manager_salary + total_office_rent + total_others
                self.summary_labels["Total Current Bill"].setText(f"₹ {total_current_bill:,.2f}")
                self.summary_labels["Manager Salary"].setText(f"₹ {total_manager_salary:,.2f}")
                self.summary_labels["Office Rent"].setText(f"₹ {total_office_rent:,.2f}")
                self.summary_labels["Others"].setText(f"₹ {total_others:,.2f}")
                self.summary_labels["Grand Total"].setText(f"₹ {grand_total:,.2f}")

                # Fetch data from other pages/databases
                trip_expense, trip_profit = self.get_trip_totals_from_db()
                vehicle_expense = self.get_vehicle_expense_from_db()

                # Calculate Business Summary Grand Total
                # Formula: Grand Total = Trip Profit - (Vehicle Expense + Office Expense)
                business_grand_total = trip_profit - (vehicle_expense + total_office_expenses)

                # Update Business Summary
                self.business_summary_labels["Trip Expenses"].setText(f"₹ {trip_expense:,.2f}")
                self.business_summary_labels["Trip Profits"].setText(f"₹ {trip_profit:,.2f}")
                self.business_summary_labels["Vehicle Expenses"].setText(f"₹ {vehicle_expense:,.2f}")
                self.business_summary_labels["Office Expenses"].setText(f"₹ {total_office_expenses:,.2f}")
                self.business_summary_labels["Grand Total"].setText(f"₹ {business_grand_total:,.2f}")
            except Exception as e:
                print(f"Error in refresh_totals: {e}")
                # Set default values on error
                for label in self.summary_labels.values():
                    label.setText("₹ 0.00")
                for label in self.business_summary_labels.values():
                    label.setText("₹ 0.00")

    def go_back_home(self):
            if self.back_callback:
                self.back_callback()


# Update MonthWidget class to integrate with database saving/deleting

class MonthWidget(QGroupBox):
    def __init__(self, month_title, notify_totals_changed, parent=None):
        super().__init__(parent)
        self.notify_totals_changed = notify_totals_changed
        self.parent_page = parent  # Store reference to parent page
        self.db_id = None  # Will be set when loaded from DB

        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold; color: #003366;
                border: 1.5px solid #336699;
                border-radius: 8px;
                margin-top: 8px;
                padding: 8px;
                background: #fcfcfc;
            }
        """)

        self.outer_layout = QVBoxLayout(self)
        self.outer_layout.setSpacing(8)

        header_hbox = QHBoxLayout()

        self.month_edit = QLineEdit()
        self.month_edit.setText(month_title)
        self.month_edit.setFont(QFont("Arial Black", 14))
        self.month_edit.setStyleSheet("color:#003366;")
        self.month_edit.setMaximumWidth(220)
        self.month_edit.textChanged.connect(self.update_groupbox_title)
        header_hbox.addWidget(self.month_edit)

        self.setTitle(self.month_edit.text())

        self.lbl_total = QLabel("Amount: ₹ 0.00")
        self.lbl_total.setFont(QFont("Arial", 12))
        self.lbl_total.setStyleSheet("color:#003366; padding-left: 20px;")
        header_hbox.addWidget(self.lbl_total)

        header_hbox.addStretch()

        self.btn_expand = QPushButton("Expand")
        self.btn_edit = QPushButton("Edit")
        self.btn_save = QPushButton("Save")
        self.btn_delete = QPushButton("Delete")

        for btn in (self.btn_expand, self.btn_edit, self.btn_save, self.btn_delete):
            btn.setStyleSheet("""
                QPushButton {
                    background: #336699;
                    color: white;
                    padding: 4px 10px;
                    border-radius: 6px;
                    min-width: 62px;
                    min-height: 28px;
                    font-size: 13px;
                }
                QPushButton:hover {
                    background: #294d66;
                }
            """)

        header_hbox.addWidget(self.btn_expand)
        header_hbox.addWidget(self.btn_edit)
        header_hbox.addWidget(self.btn_save)
        header_hbox.addWidget(self.btn_delete)

        self.outer_layout.addLayout(header_hbox)

        self.expand_widget = QWidget()
        self.expand_layout = QVBoxLayout(self.expand_widget)
        self.outer_layout.addWidget(self.expand_widget)

        self.fields = {}
        for label_text in ("Office Expenses", "Manager Salary", "Current Bill", "Other Expenses"):
            h_layout = QHBoxLayout()
            label = QLabel(label_text)
            label.setFont(QFont("Arial", 12))
            label.setStyleSheet("color:#003366; min-width:140px;")
            edit = QLineEdit("0")
            edit.setFont(QFont("Arial", 12))
            edit.setAlignment(Qt.AlignmentFlag.AlignRight)
            edit.setStyleSheet("""
                QLineEdit {
                    background: white;
                    color: #003366;
                    border: 1px solid #336699;
                    border-radius: 5px;
                    padding: 6px;
                    min-width: 80px;
                }
                QLineEdit:read-only {
                    background: #f0f8ff;
                }
            """)
            edit.textChanged.connect(self.recalc_total)
            h_layout.addWidget(label)
            h_layout.addWidget(edit)
            self.expand_layout.addLayout(h_layout)
            self.fields[label_text] = edit

        self.btn_expand.clicked.connect(self.toggle_expand)
        self.btn_edit.clicked.connect(self.enable_editing)
        self.btn_save.clicked.connect(self.save_changes)
        self.btn_delete.clicked.connect(self.confirm_delete)

        self.disable_editing()
        self.collapse()
        self.recalc_total()

    def update_groupbox_title(self, new_text: str):
        self.setTitle(new_text)

    def toggle_expand(self):
        if self.expand_widget.isVisible():
            self.collapse()
        else:
            self.expand()

    def expand(self):
        self.expand_widget.setVisible(True)
        self.btn_expand.setText("Collapse")

    def collapse(self):
        self.expand_widget.setVisible(False)
        self.btn_expand.setText("Expand")

    def enable_editing(self):
        self.expand()
        for edit in self.fields.values():
            edit.setReadOnly(False)

    def disable_editing(self):
        for edit in self.fields.values():
            edit.setReadOnly(True)

    def save_changes(self):
        confirmed = self.show_message_box_question("Confirm Save",
                                                  "Are you sure you want to save changes?")
        if confirmed:
            self.disable_editing()
            # Save to database
            if self.parent_page and hasattr(self.parent_page, 'save_month_to_db'):
                self.parent_page.save_month_to_db(self)
            self.notify_totals_changed()

    def confirm_delete(self):
        confirmed = self.show_message_box_question("Confirm Delete",
                                                  "Do you really want to delete this monthly record?\nThis action cannot be undone.")
        if confirmed:
            # Delete from database
            if self.parent_page and hasattr(self.parent_page, 'delete_month_from_db'):
                self.parent_page.delete_month_from_db(self)
            
            self.hide()
            self.setParent(None)
            self.deleteLater()
            self.notify_totals_changed()

    def show_message_box_question(self, title: str, message: str) -> bool:
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg.setFont(QFont("Arial", 12))
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #ffffff;
                color: #003366;
            }
            QPushButton {
                background-color: #336699;
                color: white;
                font-weight: bold;
                padding: 6px 12px;
                border-radius: 6px;
                min-width: 80px;
                min-height: 30px;
            }
            QPushButton:hover {
                background-color: #294d66;
            }
            QLabel {
                color: #003366;
            }
        """)
        result = msg.exec()
        return result == int(QMessageBox.StandardButton.Yes)

    def recalc_total(self):
        """Recalculate and display the total for this month"""
        try:
            total = 0.0
            for edit in self.fields.values():
                text = edit.text().strip()
                if not text:
                    continue
                try:
                    normalized = text.replace(",", "")
                    val = float(normalized)
                    total += val
                except Exception as e:
                    print(f"Error parsing value '{text}': {e}")
                    pass
            self.lbl_total.setText(f"Amount: ₹ {total:.2f}")
            
            # Safely call notify_totals_changed
            try:
                if self.notify_totals_changed:
                    self.notify_totals_changed()
            except Exception as e:
                print(f"Error in notify_totals_changed: {e}")
        except Exception as e:
            print(f"Error in recalc_total: {e}")
            self.lbl_total.setText(f"Amount: ₹ 0.00")

    def get_totals(self):
        """Get totals from all fields - always returns a dict, never None"""
        totals = {}
        try:
            for key, edit in self.fields.items():
                try:
                    text = edit.text().strip()
                    if not text:
                        totals[key] = 0.0
                    else:
                        totals[key] = float(text.replace(",", ""))
                except Exception:
                    totals[key] = 0.0
        except Exception as e:
            print(f"Error in get_totals: {e}")
            # Return default values for all expected fields
            totals = {
                "Office Expenses": 0.0,
                "Manager Salary": 0.0,
                "Current Bill": 0.0,
                "Other Expenses": 0.0
            }
        return totals

# --- Trip Manager Page Classes ---


class DateRangeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Custom Date Range")
        self.setMinimumSize(500, 300)
        self.resize(500, 300)
        
        # Enhanced styling for complete visibility
        self.setStyleSheet("""
            QDialog {
                background-color: white;
                color: black;
                border: 3px solid #336699;
            }
            QLabel {
                background-color: white;
                color: black;
                font-size: 18px;
                font-weight: bold;
                padding: 8px;
            }
            QDateEdit {
                background-color: white;
                color: black;
                font-size: 18px;
                font-weight: bold;
                border: 3px solid #336699;
                border-radius: 8px;
                padding: 10px;
                min-height: 35px;
            }
            QDateEdit::drop-down {
                border: 0px;
                background-color: #336699;
                width: 30px;
            }
            QDateEdit::down-arrow {
                image: none;
                border: none;
                background-color: #336699;
            }
            QCalendarWidget {
                background-color: white;
                color: black;
                selection-background-color: #336699;
                selection-color: white;
                font-size: 16px;
            }
            QCalendarWidget QWidget {
                background-color: white;
                color: black;
            }
            QCalendarWidget QToolButton {
                background-color: white;
                color: black;
                font-weight: bold;
            }
            QPushButton {
                background-color: #336699;
                color: white;
                font-weight: bold;
                font-size: 18px;
                padding: 12px 25px;
                border-radius: 10px;
                min-width: 120px;
                min-height: 45px;
                border: none;
            }
            QPushButton:hover {
                background-color: #294d66;
            }
        """)
        
        layout = QGridLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)

        label_from = QLabel("From Date:")
        label_to = QLabel("To Date:")
        
        self.date_from = QDateEdit(QDate.currentDate())
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("yyyy-MM-dd")
        
        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("yyyy-MM-dd")

        btn_ok = QPushButton("OK")
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(self.reject)

        layout.addWidget(label_from, 0, 0)
        layout.addWidget(self.date_from, 0, 1)
        layout.addWidget(label_to, 1, 0)
        layout.addWidget(self.date_to, 1, 1)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(btn_ok)
        button_layout.addWidget(btn_cancel)
        
        layout.addLayout(button_layout, 2, 0, 1, 2)

    def get_range(self):
        d_from = self.date_from.date().toPyDate()
        d_to = self.date_to.date().toPyDate()
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        return (d_from, d_to)

class TripDetailDialog(QDialog):
    """Enhanced dialog for entering or editing detailed trip data with Start/End Date fields."""
    def __init__(self, parent=None, row=None):
        super().__init__(parent)
        self.row = row
        self.setWindowTitle("Trip Details")
        self.setMinimumSize(750, 700)
        self.resize(750, 700)

        # Enhanced styling
        self.setStyleSheet("""
            QDialog {
                background-color: white;
                color: black;
                border: 2px solid #336699;
            }
            QLabel {
                background-color: white;
                color: black;
                font-weight: bold;
                font-size: 14px;
                padding: 4px;
            }
            QLineEdit, QDateEdit {
                background-color: white;
                color: black;
                border: 2px solid #336699;
                border-radius: 5px;
                padding: 6px;
                font-size: 14px;
                font-weight: bold;
                min-height: 20px;
            }
            QLineEdit:read-only, QDateEdit:read-only {
                background-color: #f0f0f0;
                color: #444444;
            }
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                font-size: 16px;
                padding: 10px 25px;
                border-radius: 8px;
                min-height: 35px;
                border: none;
            }
            QPushButton:hover {
                background: #219a52;
            }
        """)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Create scroll area for the form
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_widget = QWidget()
        layout = QGridLayout(scroll_widget)
        layout.setSpacing(8)
        layout.setContentsMargins(15, 15, 15, 15)

        # Enhanced fields list with Start Date and End Date
        fields = [
            "Driver Name", "Start Date", "End Date", "Start KM", "End KM", "KM Travelled", 
            "Total Trip Amount", "Trip Advance", "Return Balance", "Pooja", "Diesel", 
            "R.T.O & P.C", "Toll", "Driver Amount", "Driver Advance", "Driver Balance", 
            "Cleaner Amount", "Broker Amount", "Load Amount", "Unload Amount", "Others",
        ]

        self.entries = {}
        date_fields = {"Start Date", "End Date"}

        def get_float(field):
            if field not in self.entries:
                return 0.0
            widget = self.entries[field]
            if isinstance(widget, QDateEdit):
                return 0.0  # Date fields don't contribute to calculations
            return safe_float(widget.text())

        def calculate_km():
            start = get_float("Start KM")
            end = get_float("End KM")
            travelled = max(0, end - start)
            if "KM Travelled" in self.entries:
                self.entries["KM Travelled"].setText(show_int_amount(travelled))

        def calculate_driver_balance():
            subtract_fields = [
                "Pooja", "R.T.O & P.C", "Load Amount", "Unload Amount", "Others", "Cleaner Amount",
            ]
            subtract_total = sum(get_float(f) for f in subtract_fields)
            driver_amount = get_float("Driver Amount")
            driver_advance = get_float("Driver Advance")
            driver_balance = (driver_amount - driver_advance) + subtract_total
            if "Driver Balance" in self.entries:
                self.entries["Driver Balance"].setText(show_int_amount(driver_balance))

            # Update main table driver amount
            try:
                if self.row and self.row["entries"][5]:
                    self.row["entries"][5].setText(show_int_amount(driver_amount))
            except Exception:
                pass

        def update_driver_amount_from_total():
            try:
                total = safe_float(self.entries["Total Trip Amount"].text())
                driver_amt = total * 0.11
                self.entries["Driver Amount"].setText(show_int_amount(driver_amt))
            except Exception:
                self.entries["Driver Amount"].setText("0")
            calculate_driver_balance()

        # IMPORTANT: Function to automatically update Load Date when Start Date changes
        def on_start_date_changed(qdate):
            """When Start Date is changed, automatically update Load Date in main table"""
            if self.row and self.row["entries"][0]:
                # Convert QDate to display format (dd-MM-yyyy)
                date_str = qdate.toString("dd-MM-yyyy")
                self.row["entries"][0].setText(date_str)

        # Create form fields
        for i, field in enumerate(fields):
            label = QLabel(field + ":")

            if field in date_fields:
                # Create date picker widget
                entry = QDateEdit()
                entry.setCalendarPopup(True)
                entry.setDisplayFormat("dd-MM-yyyy")

                # Load existing date value
                existing_date = self.row["detail_entries"].get(field, "") if self.row else ""
                if existing_date:
                    # Try to parse existing date
                    qdate = QDate.fromString(existing_date, "dd-MM-yyyy")
                    if not qdate.isValid():
                        qdate = QDate.fromString(existing_date, "yyyy-MM-dd")
                    if qdate.isValid():
                        entry.setDate(qdate)
                    else:
                        entry.setDate(QDate.currentDate())
                else:
                    entry.setDate(QDate.currentDate())

                # Connect Start Date change to Load Date update
                if field == "Start Date":
                    entry.dateChanged.connect(on_start_date_changed)

            else:
                # Create regular text input
                entry = QLineEdit()
                entry.setText(self.row["detail_entries"].get(field, "") if self.row else "")

                if field in ("Driver Balance", "KM Travelled"):
                    entry.setReadOnly(True)

            self.entries[field] = entry
            layout.addWidget(label, i, 0)
            layout.addWidget(entry, i, 1)

        # Connect calculation events
        if "Total Trip Amount" in self.entries:
            self.entries["Total Trip Amount"].textChanged.connect(update_driver_amount_from_total)

        if "Start KM" in self.entries:
            self.entries["Start KM"].textChanged.connect(calculate_km)

        if "End KM" in self.entries:
            self.entries["End KM"].textChanged.connect(calculate_km)

        calc_fields = [
            "Driver Amount", "Driver Advance", "Pooja", "R.T.O & P.C", "Load Amount",
            "Unload Amount", "Others", "Cleaner Amount",
        ]
        for f in calc_fields:
            if f in self.entries:
                self.entries[f].textChanged.connect(calculate_driver_balance)

        # Initial calculations
        calculate_km()
        calculate_driver_balance()
        update_driver_amount_from_total()

        scroll_area.setWidget(scroll_widget)
        main_layout.addWidget(scroll_area)

        btn_save = QPushButton("Save & Close")
        btn_save.clicked.connect(self.save_and_close)
        main_layout.addWidget(btn_save, alignment=Qt.AlignmentFlag.AlignCenter)

    def save_and_close(self):
        # Save all field values
        for field, widget in self.entries.items():
            if isinstance(widget, QDateEdit):
                # Save date as string
                self.row["detail_entries"][field] = widget.date().toString("dd-MM-yyyy")
            else:
                # Save text field
                self.row["detail_entries"][field] = widget.text()

        # Update main table calculations
        try:
            total_val = safe_float(self.row["detail_entries"].get("Total Trip Amount", "0"))

            # Calculate total expenses from detail fields
            expense_fields = ["Pooja", "Diesel", "R.T.O & P.C", "Toll", "Driver Amount", 
                            "Cleaner Amount", "Broker Amount", "Load Amount", "Unload Amount", "Others"]
            total_expense = sum(safe_float(self.row["detail_entries"].get(f, "0")) for f in expense_fields)

            profit_val = total_val - total_expense

            # Update main table columns
            if self.row["entries"][4]:  # Load Amount column
                self.row["entries"][4].setText(show_int_amount(total_val))
            if self.row["entries"][6]:  # Expenses column
                self.row["entries"][6].setText(show_int_amount(total_expense))
            if self.row["entries"][7]:  # Profit column
                self.row["entries"][7].setText(show_int_amount(profit_val))
        except Exception:
            pass

        self.accept()

class DateRangeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Custom Date Range")
        layout = QGridLayout(self)


        label_from = QLabel("From:")
        label_to = QLabel("To:")
        self.date_from = QDateEdit(QDate.currentDate())
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd-MM-yyyy")  # Changed format
        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd-MM-yyyy")  # Changed format


        btn_ok = QPushButton("OK")
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(self.reject)


        layout.addWidget(label_from, 0, 0)
        layout.addWidget(self.date_from, 0, 1)
        layout.addWidget(label_to, 1, 0)
        layout.addWidget(self.date_to, 1, 1)
        layout.addWidget(btn_ok, 2, 0)
        layout.addWidget(btn_cancel, 2, 1)


    def get_range(self):
        d_from = self.date_from.date().toPyDate()
        d_to = self.date_to.date().toPyDate()
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        return (d_from, d_to)

class TripManagerPage(QMainWindow):
    """COMPLETELY FIXED Trip Manager page with proper sqlite3.Row handling."""

    def __init__(self, back_callback=None, db_manager=None):
        super().__init__()
        self.setWindowTitle("Trip Manager")
        self.setGeometry(100, 100, 1280, 780)
        self.back_callback = back_callback
        self.db = db_manager  # Store DB manager reference
        self.rows = []
        self.custom_range = (None, None)
        self.expand_dialog = None

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.central_widget.setStyleSheet(f"background-color: {NAVY}; color: white;")
        self.main_layout = QVBoxLayout(self.central_widget)

        self.create_top_bar()
        self.create_table()
        self.create_summary_box()

        # Load trips from DB if available
        if self.db is not None:
            self.load_from_db()

    def create_top_bar(self):
        top_bar = QHBoxLayout()

        self.button_back = QPushButton("Back to Home")
        self.button_back.setStyleSheet(f"background: #e67e22; color: white; font-weight: bold; font-size: 14px;")
        self.button_back.clicked.connect(self.back_to_home)
        top_bar.addWidget(self.button_back)

        self.button_new = QPushButton("NEW")
        self.button_new.setStyleSheet(f"background: #27ae60; color: white; font-weight: bold; font-size: 14px;")
        self.button_new.clicked.connect(self.add_row)
        top_bar.addWidget(self.button_new)

        self.button_download = QPushButton("DOWNLOAD")
        self.button_download.setStyleSheet(f"background: #2980b9; color: white; font-weight: bold; font-size: 14px;")
        self.button_download.clicked.connect(self.show_download_menu)
        top_bar.addWidget(self.button_download)

        self.date_option = QComboBox()
        self.date_option.addItems(
            ["Filter", "Last 1 Day", "Last Month", "Last 3 Months",
             "Last 6 Months", "Last Year", "Custom..."]
        )
        self.date_option.setStyleSheet("background-color: white; color: black; font-weight: normal; font-size: 13px;")
        self.date_option.currentTextChanged.connect(self.set_date_option)
        top_bar.addWidget(self.date_option)

        label_vehicle = QLabel("Vehicle")
        label_vehicle.setStyleSheet(f"color: white; font-weight: bold; font-size: 14px; margin-left: 8px;")
        top_bar.addWidget(label_vehicle)

        self.vehicle_var = QLineEdit()
        self.vehicle_var.setStyleSheet("background-color: white; color: black; font-size: 14px;")
        top_bar.addWidget(self.vehicle_var)

        label_broker = QLabel("Broker Office")
        label_broker.setStyleSheet(f"color: white; font-weight: bold; font-size: 14px; margin-left: 8px;")
        top_bar.addWidget(label_broker)

        self.brokeroffice_var = QLineEdit()
        self.brokeroffice_var.setPlaceholderText("Search Broker Office")
        self.brokeroffice_var.setStyleSheet("background-color: white; color: black; font-size: 14px;")
        top_bar.addWidget(self.brokeroffice_var)

        label_driver = QLabel("Driver Name")
        label_driver.setStyleSheet(f"color: white; font-weight: bold; font-size: 14px; margin-left: 8px;")
        top_bar.addWidget(label_driver)

        self.driver_var = QLineEdit()
        self.driver_var.setPlaceholderText("Search Driver Name")
        self.driver_var.setStyleSheet("background-color: white; color: black; font-size: 14px;")
        top_bar.addWidget(self.driver_var)

        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "Paid", "Unpaid"])
        self.status_filter.setStyleSheet("background-color: white; color: black; font-weight: normal; font-size: 13px; margin-left: 8px;")
        self.status_filter.currentTextChanged.connect(self.search)
        top_bar.addWidget(self.status_filter)

        self.button_search = QPushButton("SEARCH")
        self.button_search.setStyleSheet("background: #c0392b; color: white; font-weight: bold; font-size: 14px;")
        self.button_search.clicked.connect(self.search)
        top_bar.addWidget(self.button_search)

        self.button_reset = QPushButton("RESET")
        self.button_reset.setStyleSheet("background: #7f8c8d; color: white; font-weight: bold; font-size: 14px;")
        self.button_reset.clicked.connect(self.reset)
        top_bar.addWidget(self.button_reset)

        top_frame = QFrame()
        top_frame.setStyleSheet(f"background: {NAVY_DARK}; color: white;")
        top_frame.setLayout(top_bar)
        self.main_layout.addWidget(top_frame)

    def show_download_menu(self):
        menu = QMenu(self)
        menu.setStyleSheet("""
        QMenu {
            background-color: white;
            color: black;
            border: 1px solid #ccc;
        }
        QMenu::item {
            padding: 8px 20px;
        }
        QMenu::item:selected {
            background-color: #2980b9;
            color: white;
        }
        """)
        pdf_action = QAction("Download PDF", self)
        pdf_action.triggered.connect(self.download_pdf)
        menu.addAction(pdf_action)

        excel_action = QAction("Download Excel", self)
        excel_action.triggered.connect(self.download_excel)
        menu.addAction(excel_action)

        menu.exec(self.button_download.mapToGlobal(self.button_download.rect().bottomLeft()))

    def back_to_home(self):
        if self.back_callback:
            self.back_callback()

    def create_table(self):
        headers = [
            "Load Date", "Location From - To", "Vehicle No", "Broker Office",
            "Load Amount", "Driver Amount", "Expenses", "Profit",
            "Status", "Expand", "Delete"
        ]

        self.table = QTableWidget(0, len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            f"QTableWidget {{background-color: {WHITE}; color: black; font-size: 14px;}}"
            f"QHeaderView::section {{background-color: {NAVY}; color: white; font-weight: bold;}}"
            f"QTableWidget::item:selected {{background-color: {SELECT_ROW_COLOR}; color: black;}}"
            f"QTableWidget::item:alternate {{background-color: {LIGHT_GRAY};}}"
        )

        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        header = self.table.horizontalHeader()
        for i in range(self.table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)

        self.main_layout.addWidget(self.table)

    def create_summary_box(self):
        summary_layout = QHBoxLayout()

        self.total_driver_amount_var = QLabel("0")
        self.total_profit_var = QLabel("0")
        self.total_expense_var = QLabel("0")
        self.total_sum_var = QLabel("0")
        self.total_unpaid_var = QLabel("0")

        def mk_label(text, label):
            lab = QLabel(text)
            lab.setStyleSheet("color: white; font-weight: bold; font-size: 14px;")
            label.setStyleSheet("background-color: white; color: black; padding: 5px 12px; font-size: 14px; min-width: 100px;")
            summary_layout.addWidget(lab)
            summary_layout.addWidget(label)

        mk_label("Total Driver Amount:", self.total_driver_amount_var)
        mk_label("Total Profit:", self.total_profit_var)
        mk_label("Total Expenses:", self.total_expense_var)
        mk_label("Total Load Amount:", self.total_sum_var)
        mk_label("Total Unpaid Amount:", self.total_unpaid_var)

        summary_frame = QFrame()
        summary_frame.setStyleSheet(f"background: {NAVY_DARK}; color: white; padding: 10px;")
        summary_frame.setLayout(summary_layout)
        self.main_layout.addWidget(summary_frame)

    # -------------------- DATABASE INTEGRATION METHODS (FIXED FOR sqlite3.Row) --------------------

    def row_to_dict(self, row):
        """Convert sqlite3.Row to dictionary for easier handling."""
        if row is None:
            return {}
        return dict(row)

    def load_from_db(self):
        """Load all trips from database into the table."""
        if not self.db:
            return

        # Clear existing data
        self.table.setRowCount(0)
        self.rows.clear()

        try:
            # Load trips from DB and convert to dictionaries
            trips_rows = self.db.load_trips()
            trips = [self.row_to_dict(trip) for trip in trips_rows]

            for trip in trips:
                self.insert_trip_row_from_db(trip)

            self.update_summary()
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load trips: {str(e)}")

    def insert_trip_row_from_db(self, trip_dict):
        """Insert a single trip row from database dictionary into the table."""
        row = 0  # Insert at first row instead of last
        self.table.insertRow(0)

        # Parse detail_json if it exists
        detail_entries = {}
        detail_json = trip_dict.get('detail_json')
        if detail_json:
            try:
                detail_entries = json.loads(detail_json)
            except:
                detail_entries = {}

        # Create table items
        items = []

        # Column 0: Date
        date_value = trip_dict.get('date', '')
        date_item = QTableWidgetItem(format_date_for_display(date_value))
        date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        date_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled )
        self.table.setItem(row, 0, date_item)
        items.append(date_item)

        # Column 1: Location
        location_value = trip_dict.get('location_from_to', '') or ""
        loc_item = QTableWidgetItem(str(location_value))
        loc_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        loc_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 1, loc_item)
        items.append(loc_item)

        # Column 2: Vehicle No
        vehicle_value = trip_dict.get('vehicle_no', '') or ""
        veh_item = QTableWidgetItem(str(vehicle_value))
        veh_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        veh_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 2, veh_item)
        items.append(veh_item)

        # Column 3: Broker Office
        broker_value = trip_dict.get('broker_office', '') or ""
        broker_item = QTableWidgetItem(str(broker_value))
        broker_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        broker_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 3, broker_item)
        items.append(broker_item)

        # Column 4: Load Amount (Total)
        total_value = trip_dict.get('total', 0) or 0
        load_item = QTableWidgetItem(show_int_amount(total_value))
        load_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        load_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 4, load_item)
        items.append(load_item)

        # Column 5: Driver Amount
        driver_value = trip_dict.get('driver_amount', 0) or 0 # Use driver_amount from DB
        driver_item = QTableWidgetItem(show_int_amount(driver_value))
        driver_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        driver_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 5, driver_item)
        items.append(driver_item)

        # Column 6: Expenses
        expense_value = trip_dict.get('expense', 0) or 0
        expense_item = QTableWidgetItem(show_int_amount(expense_value))
        expense_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        expense_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 6, expense_item)
        items.append(expense_item)

        # Column 7: Profit
        profit_value = trip_dict.get('profit', 0) or 0
        profit_item = QTableWidgetItem(show_int_amount(profit_value))
        profit_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        profit_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 7, profit_item)
        items.append(profit_item)

        # Column 8: Status
        status_value = trip_dict.get('status', 'Unpaid') or "Unpaid"
        status_item = QTableWidgetItem(str(status_value))
        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        status_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 8, status_item)
        items.append(status_item)

        # Add buttons for columns 9 and 10
        expand_btn = QPushButton("Expand")
        expand_btn.setStyleSheet("background: #3498db; color: white; font-weight: bold;")
        expand_btn.clicked.connect(lambda checked=False, r=row: self.expand_clicked(r))
        self.table.setCellWidget(row, 9, expand_btn)

        delete_btn = QPushButton("X")
        delete_btn.setStyleSheet("background: #e74c3c; color: white; font-weight: bold;")
        delete_btn.clicked.connect(lambda checked=False, r=row: self.delete_clicked(r))
        self.table.setCellWidget(row, 10, delete_btn)

        # Store row data with DB ID for future reference
        row_data = {
            "db_id": trip_dict.get('id'),
            "entries": items,  # Store QTableWidgetItems
            "detail_entries": detail_entries,
            "expand_btn": expand_btn,
            "delete_btn": delete_btn,
        }
        self.rows.insert(0, row_data)  # Insert at first position instead of append

        # Update button connections for all existing rows (their indices shifted by 1)
        for i in range(1, len(self.rows)):
            if 'expand_btn' in self.rows[i] and 'delete_btn' in self.rows[i]:
                # Disconnect old connections
                try:
                    self.rows[i]['expand_btn'].clicked.disconnect()
                    self.rows[i]['delete_btn'].clicked.disconnect()
                except:
                    pass
                # Reconnect with updated index
                self.rows[i]['expand_btn'].clicked.connect(lambda checked=False, r=i: self.expand_clicked(r))
                self.rows[i]['delete_btn'].clicked.connect(lambda checked=False, r=i: self.delete_clicked(r))

    def add_row(self):
        """Add a new trip row and save it to database immediately."""
        if not self.db:
            QMessageBox.warning(self, "No Database", "Database connection not available.")
            return

        # Create new trip record in database
        current_date = datetime.now().strftime("%Y-%m-%d")
        trip_data = (
            current_date,  # date
            "",           # vehicle_no
            "",           # location_from_to
            "",           # broker_office
            0.0,          # driver_amount
            0.0,          # profit
            0.0,          # expense
            0.0,          # total
            "Unpaid",     # status
            json.dumps({}) # detail_json
        )

        try:
            new_id = self.db.save_trip(trip_data)
            # Fetch the newly created record and add it to the table
            new_trip_row = self.db.conn.execute("SELECT * FROM trips WHERE id=?", (new_id,)).fetchone()
            if new_trip_row:
                new_trip_dict = self.row_to_dict(new_trip_row)
                self.insert_trip_row_from_db(new_trip_dict)
                self.update_summary()
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Failed to add new trip: {str(e)}")

    def expand_clicked(self, row_index):
        """Open trip details dialog for editing."""
        if row_index >= len(self.rows):
            return

        row_data = self.rows[row_index]
        dlg = TripDetailDialog(self, row=row_data)

        if dlg.exec():
            # Update database with new values
            if self.db and 'db_id' in row_data:
                self.save_trip_to_db(row_data)
                # Refresh the row display
                self.refresh_row_from_db(row_index)
                self.update_summary()

    def delete_clicked(self, row_index):
        """Delete a trip row."""
        if row_index >= len(self.rows):
            return

        reply = QMessageBox.question(
            self, "Confirm Delete", 
            "Are you sure you want to delete this trip?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            row_data = self.rows[row_index]

            # Delete from database
            if self.db and 'db_id' in row_data:
                try:
                    self.db.delete_trip(row_data['db_id'])
                except Exception as e:
                    QMessageBox.critical(self, "Database Error", f"Failed to delete trip: {str(e)}")
                    return

            # Remove from table and rows list
            self.table.removeRow(row_index)
            del self.rows[row_index]

            # Update row indices for remaining rows
            for i in range(row_index, len(self.rows)):
                # Update button connections
                self.rows[i]['expand_btn'].clicked.disconnect()
                self.rows[i]['delete_btn'].clicked.disconnect()
                self.rows[i]['expand_btn'].clicked.connect(lambda checked=False, r=i: self.expand_clicked(r))
                self.rows[i]['delete_btn'].clicked.connect(lambda checked=False, r=i: self.delete_clicked(r))

            self.update_summary()

    def save_trip_to_db(self, row_data):
        """Save trip data to database."""
        if not self.db or 'db_id' not in row_data:
            return

        entries = row_data['entries']
        detail_entries = row_data['detail_entries']

        # Calculate totals from detail entries
        total_trip = safe_float(detail_entries.get('Total Trip Amount', 0))

        expense_fields = ["Pooja", "Diesel", "R.T.O & P.C", "Toll", "Driver Amount", 
                         "Cleaner Amount", "Broker Amount", "Load Amount", "Unload Amount", "Others"]
        total_expense = sum(safe_float(detail_entries.get(f, 0)) for f in expense_fields)

        driver_amount = safe_float(detail_entries.get('Driver Balance', 0))
        profit = total_trip - total_expense

        # Get basic info from table entries
        date_str = entries[0].text() if entries[0] else ""
        # Convert from display format (dd-mm-yyyy) to storage format (yyyy-mm-dd)
        try:
            if date_str:
                date_obj = datetime.strptime(date_str, "%d-%m-%Y")
                date_str = date_obj.strftime("%Y-%m-%d")
        except:
            date_str = datetime.now().strftime("%Y-%m-%d")

        location = entries[1].text() if entries[1] else ""
        vehicle_no = entries[2].text() if entries[2] else ""
        broker_office = entries[3].text() if entries[3] else ""

        # Compute payment completion
        total_trip = safe_float(total_trip)
        trip_advance = safe_float(detail_entries.get('Trip Advance', 0))
        return_balance = safe_float(detail_entries.get('Return Balance', 0))
        broker_amount = safe_float(detail_entries.get('Broker Amount', 0))

        # Determine payment status
        difference = total_trip - (trip_advance + return_balance + broker_amount)

        if abs(difference) < 0.01:
            status = "PAID"
        elif difference > 0:
            status = "UNPAID"
        else:
            QMessageBox.warning(self, "Mismatch", "Check the inputs — total exceeds trip amount.")
            return()
        
        # Update database
        update_data = (
            date_str,
            vehicle_no, 
            location,
            broker_office,
            driver_amount,
            profit,
            total_expense,
            total_trip,
            status,
            json.dumps(detail_entries)
        )

        try:
            self.db.update_trip(row_data['db_id'], update_data)
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Failed to save trip: {str(e)}")

    def refresh_row_from_db(self, row_index):
        """Refresh a single row from database."""
        if not self.db or row_index >= len(self.rows):
            return

        row_data = self.rows[row_index]
        if 'db_id' not in row_data:
            return

        # Fetch updated data from database
        trip_row = self.db.conn.execute("SELECT * FROM trips WHERE id=?", (row_data['db_id'],)).fetchone()
        if not trip_row:
            return

        trip_dict = self.row_to_dict(trip_row)

        # Update table items
        entries = row_data['entries']

        if entries[0]:  # Date
            entries[0].setText(format_date_for_display(trip_dict.get('date', '')))
        if entries[1]:  # Location
            entries[1].setText(str(trip_dict.get('location_from_to', '') or ""))
        if entries[2]:  # Vehicle
            entries[2].setText(str(trip_dict.get('vehicle_no', '') or ""))
        if entries[3]:  # Broker
            entries[3].setText(str(trip_dict.get('broker_office', '') or ""))
        if entries[4]:  # Load Amount
            entries[4].setText(show_int_amount(trip_dict.get('total', 0) or 0))
        if entries[5]:  # Driver Amount
            entries[5].setText(show_int_amount(trip_dict.get('driver_amount', 0) or 0))
        if entries[6]:  # Expenses
            entries[6].setText(show_int_amount(trip_dict.get('expense', 0) or 0))
        if entries[7]:  # Profit
            entries[7].setText(show_int_amount(trip_dict.get('profit', 0) or 0))
        if entries[8]:  # Status
            entries[8].setText(str(trip_dict.get('status', 'Unpaid') or "Unpaid"))

    # -------------------- OTHER METHODS --------------------

    def update_summary(self):
        """Update summary totals."""
        total_driver = 0
        total_profit = 0
        total_expense = 0
        total_sum = 0
        total_unpaid = 0

        for row_data in self.rows:
            entries = row_data['entries']
            detail_entries = row_data['detail_entries']

            # Get values from table items
            load_amount = safe_float(entries[4].text() if entries[4] else "0")
            driver_amount = safe_float(entries[5].text() if entries[5] else "0")
            expense_amount = safe_float(entries[6].text() if entries[6] else "0")
            profit_amount = safe_float(entries[7].text() if entries[7] else "0")
            status = entries[8].text().lower() if entries[8] else "unpaid"

            total_sum += load_amount
            total_driver += driver_amount
            total_expense += expense_amount
            total_profit += profit_amount

            # Calculate unpaid amount
            if status != "paid":
                trip_advance = safe_float(detail_entries.get("Trip Advance", "0"))
                return_balance = safe_float(detail_entries.get("Return Balance", "0"))
                broker_amount = safe_float(detail_entries.get("Broker Amount", "0"))
                unpaid = max(0, (load_amount - trip_advance) - broker_amount - return_balance)
                total_unpaid += unpaid
        # Update labels
        self.total_sum_var.setText(show_int_amount(total_sum))
        self.total_driver_amount_var.setText(show_int_amount(total_driver))
        self.total_expense_var.setText(show_int_amount(total_expense))
        self.total_profit_var.setText(show_int_amount(total_profit))
        self.total_unpaid_var.setText(show_int_amount(total_unpaid))

    def reset(self):
        """Reset filters and reload from database."""
        # Clear filter controls
        self.vehicle_var.clear()
        self.brokeroffice_var.clear()
        self.driver_var.clear()
        self.date_option.setCurrentIndex(0)
        self.status_filter.setCurrentIndex(0)

        # Reload from database
        self.load_from_db()

    def search(self):
        """Apply search filters."""
        vehicle_filter = self.vehicle_var.text().lower().strip()
        broker_filter = self.brokeroffice_var.text().lower().strip()
        driver_filter = self.driver_var.text().lower().strip()
        status_filter = self.status_filter.currentText()

        for i in range(self.table.rowCount()):
            show_row = True

            # Vehicle filter
            if vehicle_filter:
                vehicle_text = self.table.item(i, 2).text().lower() if self.table.item(i, 2) else ""
                if vehicle_filter not in vehicle_text:
                    show_row = False

            # Broker filter
            if broker_filter:
                broker_text = self.table.item(i, 3).text().lower() if self.table.item(i, 3) else ""
                if broker_filter not in broker_text:
                    show_row = False

            # Status filter
            if status_filter != "All":
                status_text = self.table.item(i, 8).text().lower() if self.table.item(i, 8) else "unpaid"
                if status_filter.lower() != status_text:
                    show_row = False

            # Driver filter (from detail entries)
            if driver_filter and i < len(self.rows):
                detail_entries = self.rows[i]['detail_entries']
                driver_name = detail_entries.get('Driver Name', '').lower()
                if driver_filter not in driver_name:
                    show_row = False

            self.table.setRowHidden(i, not show_row)

    def set_date_option(self, option):
        """Handle date filter option changes."""
        if option == "Custom...":
            dlg = DateRangeDialog(self)
            if dlg.exec():
                start_date, end_date = dlg.get_range()
                self.custom_range = (start_date, end_date)
                self.apply_date_filter()
        elif option != "Filter":
            self.apply_date_filter(option)

    def apply_date_filter(self, option=None):
        """Apply date-based filtering."""
        if not option:
            option = self.date_option.currentText()

        today = datetime.now().date()
        start_date = None
        end_date = today

        if option == "Last 1 Day":
            start_date = today - timedelta(days=1)
        elif option == "Last Month":
            start_date = today - timedelta(days=30)
        elif option == "Last 3 Months":
            start_date = today - timedelta(days=90)
        elif option == "Last 6 Months":
            start_date = today - timedelta(days=180)
        elif option == "Last Year":
            start_date = today - timedelta(days=365)
        elif option == "Custom...":
            if self.custom_range[0]:
                start_date, end_date = self.custom_range

        if start_date:
            for i in range(self.table.rowCount()):
                date_item = self.table.item(i, 0)
                if date_item:
                    row_date = parse_date_from_display(date_item.text())
                    if row_date:
                        show_row = start_date <= row_date <= end_date
                        self.table.setRowHidden(i, not show_row)

    def download_pdf(self):
        """Export currently visible trips to PDF using reportlab."""
        try:
            path, _ = QFileDialog.getSaveFileName(self, "Save PDF", "", "PDF Files (*.pdf)")
            if not path:
                return
            if not path.lower().endswith('.pdf'):
                path += '.pdf'

            try:
                from reportlab.lib.pagesizes import landscape, letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
            except ImportError:
                QMessageBox.warning(self, "Missing Library",
                    "ReportLab package is required for PDF export.\n\n"
                    "Install it using: pip install reportlab")
                return

            # Collect data from visible table rows with calculated status amounts
            headers = ["Load Date", "Location From-To", "Vehicle No", "Broker Office",
                      "Load Amt", "Driver Amt", "Expenses", "Profit", "Status"]
            data = [headers]

            # Initialize totals
            total_load_amt = 0
            total_driver_amt = 0
            total_expenses = 0
            total_profit = 0
            total_status = 0

            for row in range(self.table.rowCount()):
                if self.table.isRowHidden(row):
                    continue
                row_data = []
                for col in range(8):  # First 8 columns (excluding status)
                    item = self.table.item(row, col)
                    cell_text = item.text() if item else ""
                    row_data.append(cell_text)

                    # Add to totals for numeric columns
                    if col == 4:  # Load Amt
                        total_load_amt += _safefloat(cell_text)
                    elif col == 5:  # Driver Amt
                        total_driver_amt += _safefloat(cell_text)
                    elif col == 6:  # Expenses
                        total_expenses += _safefloat(cell_text)
                    elif col == 7:  # Profit
                        total_profit += _safefloat(cell_text)

                # Calculate status amount instead of just showing paid/unpaid
                if row < len(self.rows):
                    row_info = self.rows[row]
                    entries = row_info['entries']
                    detail_entries = row_info['detail_entries']

                    status = entries[8].text().lower() if entries[8] else "unpaid"
                    load_amount = _safefloat(entries[4].text() if entries[4] else "0")

                    if status == "paid":
                        status_display = "0"
                        status_amount = 0
                    else:
                        # Calculate unpaid amount
                        trip_advance = _safefloat(detail_entries.get("Trip Advance", "0"))
                        return_balance = _safefloat(detail_entries.get("Return Balance", "0"))
                        broker_amount = _safefloat(detail_entries.get("Broker Amount", "0"))
                        unpaid_amount = max(0, (load_amount - trip_advance) - broker_amount - return_balance)
                        status_display = str(int(unpaid_amount)) if unpaid_amount == int(unpaid_amount) else f"{unpaid_amount:.2f}"
                        status_amount = unpaid_amount
                else:
                    status_display = "0"
                    status_amount = 0

                total_status += status_amount
                row_data.append(status_display)
                data.append(row_data)

            if len(data) <= 1:  # Only headers
                QMessageBox.information(self, "No Data", "No trip data to export.")
                return

            # Add totals row
            totals_row = [
                "TOTAL",  # Load Date
                "",       # Location From-To
                "",       # Vehicle No
                "",       # Broker Office
                str(int(total_load_amt)) if total_load_amt == int(total_load_amt) else f"{total_load_amt:.2f}",  # Load Amt
                str(int(total_driver_amt)) if total_driver_amt == int(total_driver_amt) else f"{total_driver_amt:.2f}",  # Driver Amt
                str(int(total_expenses)) if total_expenses == int(total_expenses) else f"{total_expenses:.2f}",  # Expenses
                str(int(total_profit)) if total_profit == int(total_profit) else f"{total_profit:.2f}",  # Profit
                str(int(total_status)) if total_status == int(total_status) else f"{total_status:.2f}"   # Status
            ]
            data.append(totals_row)

            # Create PDF document
            doc = SimpleDocTemplate(
                path,
                pagesize=landscape(letter),
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch
            )

            # Create table
            table = Table(data)

            # Style the table
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                # Style the totals row
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
            ]

            table.setStyle(TableStyle(style))

            # Build PDF
            doc.build([table])

            QMessageBox.information(self, "Success", f"PDF exported successfully to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF:\n{str(e)}")

    
    def download_excel(self):
        """Download trips as Excel with totals."""
        try:
            path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
            if not path:
                return
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'

            # Collect visible rows data with calculated status amounts
            data = []
            headers = ["Load Date", "Location From - To", "Vehicle No", "Broker Office",
                      "Load Amt", "Driver Amt", "Expenses", "Profit", "Status"]

            # Initialize totals
            total_load_amt = 0
            total_driver_amt = 0
            total_expenses = 0
            total_profit = 0
            total_status = 0

            for row in range(self.table.rowCount()):
                if self.table.isRowHidden(row):
                    continue

                row_data = []
                for col in range(8):  # First 8 columns (excluding status)
                    item = self.table.item(row, col)
                    cell_text = item.text() if item else ""
                    row_data.append(cell_text)

                    # Add to totals for numeric columns
                    if col == 4:  # Load Amt
                        total_load_amt += _safefloat(cell_text)
                    elif col == 5:  # Driver Amt
                        total_driver_amt += _safefloat(cell_text)
                    elif col == 6:  # Expenses
                        total_expenses += _safefloat(cell_text)
                    elif col == 7:  # Profit
                        total_profit += _safefloat(cell_text)

                # Calculate status amount
                if row < len(self.rows):
                    row_info = self.rows[row]
                    entries = row_info['entries']
                    detail_entries = row_info['detail_entries']

                    status = entries[8].text().lower() if entries[8] else "unpaid"
                    load_amount = _safefloat(entries[4].text() if entries[4] else "0")

                    if status == "paid":
                        status_amount = 0
                    else:
                        trip_advance = _safefloat(detail_entries.get("Trip Advance", "0"))
                        return_balance = _safefloat(detail_entries.get("Return Balance", "0"))
                        broker_amount = _safefloat(detail_entries.get("Broker Amount", "0"))
                        unpaid_amount = max(0, (load_amount - trip_advance) - broker_amount - return_balance)
                        status_amount = unpaid_amount
                else:
                    status_amount = 0

                total_status += status_amount
                row_data.append(status_amount)
                data.append(row_data)

            if not data:
                QMessageBox.information(self, "No Data", "No trip data to export.")
                return

            # Try openpyxl first (preferred for Excel formatting)
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Trip Data"

                # Add headers with formatting
                for c, header in enumerate(headers, 1):
                    cell = ws.cell(1, c, header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # Add data rows
                for r, row_data in enumerate(data, start=2):
                    for c, value in enumerate(row_data, 1):
                        ws.cell(r, c, value)

                # Add totals row
                totals_row_num = len(data) + 2
                ws.cell(totals_row_num, 1, "TOTAL").font = Font(bold=True)
                ws.cell(totals_row_num, 5, total_load_amt).font = Font(bold=True)    # Load Amt
                ws.cell(totals_row_num, 6, total_driver_amt).font = Font(bold=True)  # Driver Amt
                ws.cell(totals_row_num, 7, total_expenses).font = Font(bold=True)    # Expenses
                ws.cell(totals_row_num, 8, total_profit).font = Font(bold=True)      # Profit
                ws.cell(totals_row_num, 9, total_status).font = Font(bold=True)      # Status

                wb.save(path)

            except ImportError:
                # Fallback to pandas
                try:
                    import pandas as pd

                    # Create DataFrame
                    df = pd.DataFrame(data, columns=headers)

                    # Add totals row
                    totals_row = {
                        "Load Date": "TOTAL",
                        "Location From - To": "",
                        "Vehicle No": "",
                        "Broker Office": "",
                        "Load Amt": total_load_amt,
                        "Driver Amt": total_driver_amt,
                        "Expenses": total_expenses,
                        "Profit": total_profit,
                        "Status": total_status
                    }

                    df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
                    df.to_excel(path, index=False, sheet_name="Trip Data")

                except ImportError:
                    QMessageBox.warning(self, "Missing Library",
                        "Either openpyxl or pandas package is required for Excel export.\n\n"
                        "Install using: pip install openpyxl\nor: pip install pandas openpyxl")
                    return

            QMessageBox.information(self, "Success", f"Excel file exported successfully to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n{str(e)}")

    
class MainWindow(QMainWindow):
    def __init__(self, db_manager=None):
        super().__init__()
        self.db = db_manager
        self.setWindowTitle("KTS Transport")
        self.showMaximized()
        self.setStyleSheet("background-color: white;")
        self.show_home_page()
    
    def show_home_page(self):
        """Display the main home page with navigation cards"""
        try:
            central_widget = QWidget()
            self.setCentralWidget(central_widget)
            
            main_layout = QVBoxLayout(central_widget)
            main_layout.setContentsMargins(40, 40, 40, 40)
            main_layout.setSpacing(40)
            
            # Header
            header_frame = QFrame()
            header_frame.setStyleSheet("background-color: #e9f7ff; border-radius: 12px;")
            
            header_layout = QHBoxLayout(header_frame)
            header_layout.setContentsMargins(30, 10, 30, 10)
            header_layout.setSpacing(20)
            header_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            title_container = QVBoxLayout()
            title_container.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            title_label = QLabel("KTS TRANSPORT")
            title_label.setFont(QFont("Arial Black", 36, QFont.Weight.Bold))
            title_label.setStyleSheet("color: #007BFF;")
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_container.addWidget(title_label)
            
            desc_label = QLabel("134, G NVK COMPLEX, SALEM ROAD, NAMAKKAL - 637001\n638244-7660")
            desc_label.setFont(QFont("Arial", 14))
            desc_label.setStyleSheet("color: #333333;")
            desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_container.addWidget(desc_label)
            
            header_layout.addLayout(title_container)
            main_layout.addWidget(header_frame)
            
            # Cards container
            cards_container = QWidget()
            cards_layout = QVBoxLayout(cards_container)
            cards_layout.setContentsMargins(0, 0, 0, 0)
            cards_layout.setSpacing(40)
            
            # Top row
            top_row = QHBoxLayout()
            top_row.setSpacing(60)
            top_row.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            vehicle_driver_card = self.make_card("👮‍♂️", "Vehicle and Driver", "Driver and Vehicle Management", 
                                                clickable=True, click_handler="vehicledriver")
            vehicle_card = self.make_card("🚚", "Vehicle", "Fleet and Vehicle Expenses", 
                                         clickable=True, click_handler="vehicle")
            
            top_row.addWidget(vehicle_driver_card)
            top_row.addWidget(vehicle_card)
            cards_layout.addLayout(top_row)
            
            # Bottom row
            bottom_row = QHBoxLayout()
            bottom_row.setSpacing(60)
            bottom_row.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            trip_card = self.make_card("🗺️", "Trip", "Trip Planning & Tracking", 
                                      clickable=True, click_handler="trip")
            office_card = self.make_card("🏢", "Office", "Office Operations", 
                                        clickable=True, click_handler="office")
            
            bottom_row.addWidget(trip_card)
            bottom_row.addWidget(office_card)
            cards_layout.addLayout(bottom_row)
            
            main_layout.addWidget(cards_container)
        except Exception as e:
            print(f"Error showing home page: {e}")
            QMessageBox.critical(self, "Error", f"Failed to show home page: {str(e)}")
    
    def make_card(self, icon, title, desc, clickable=False, click_handler=None):
        """Create a navigation card with icon, title, and description"""
        try:
            frame = QFrame()
            frame.setStyleSheet("""
            QFrame {
                background-color: #e9f2ff;
                border-radius: 12px;
            }
            QFrame:hover {
                background-color: #d4ecff;
            }
            """)
            frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
            frame.setMinimumSize(700, 250)
            
            vbox = QVBoxLayout(frame)
            vbox.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.setSpacing(10)
            
            lbl_icon = QLabel(icon)
            lbl_icon.setFont(QFont("Segoe UI Emoji", 48))
            lbl_icon.setStyleSheet("color: #007BFF;")
            lbl_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.addWidget(lbl_icon)
            
            lbl_title = QLabel(title)
            lbl_title.setFont(QFont("Arial", 20, QFont.Weight.Bold))
            lbl_title.setStyleSheet("""
            color: #003366;
            background-color: #d8ecff;
            padding: 6px 16px;
            border-radius: 10px;
            """)
            lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.addWidget(lbl_title)
            
            lbl_desc = QLabel(desc)
            lbl_desc.setFont(QFont("Arial", 13))
            lbl_desc.setStyleSheet("color: #444444;")
            lbl_desc.setWordWrap(True)
            lbl_desc.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vbox.addWidget(lbl_desc)
            
            # Set up click handlers for navigation
            if clickable and click_handler:
                frame.setCursor(Qt.CursorShape.PointingHandCursor)
                if click_handler == "office":
                    frame.mousePressEvent = lambda event: self.open_office_expenses()
                elif click_handler == "trip":
                    frame.mousePressEvent = lambda event: self.open_trip_manager()
                elif click_handler == "vehicledriver":
                    frame.mousePressEvent = lambda event: self.open_vehicle_driver()
                elif click_handler == "vehicle":
                    frame.mousePressEvent = lambda event: self.open_vehicle_expenses()
            
            return frame
        except Exception as e:
            print(f"Error creating card: {e}")
            return QFrame()  # Return empty frame as fallback
    
    def open_office_expenses(self):
        """Navigate to Office Expenses page"""
        try:
            self.setCentralWidget(OfficeExpensePage(self.show_home_page, self.db))
        except Exception as e:
            print(f"Error opening Office Expenses: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open Office Expenses: {str(e)}")
    
    def open_trip_manager(self):
        """Navigate to Trip Manager page"""
        try:
            self.setCentralWidget(TripManagerPage(self.show_home_page, self.db))
        except Exception as e:
            print(f"Error opening Trip Manager: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open Trip Manager: {str(e)}")
    
    def open_vehicle_driver(self):
        """Navigate to Vehicle & Driver Management page"""
        try:
            self.setCentralWidget(VehicleDriverPage(self.show_home_page, self.db))
        except Exception as e:
            print(f"Error opening Vehicle Driver: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open Vehicle Driver: {str(e)}")
    
    def open_vehicle_expenses(self):
        """Navigate to Vehicle Expenses page"""
        try:
            print("Opening Vehicle Expenses page...")
            expense_page = VehicleExpensePage(self.show_home_page, self.db)
            self.setCentralWidget(expense_page)
            print("Vehicle Expenses page opened successfully")
        except Exception as e:
            print(f"Error opening Vehicle Expenses: {e}")
            QMessageBox.critical(self, "Error", f"Failed to open Vehicle Expenses: {str(e)}")

# ----------------------------------------------------------------------
# Application Entry Point
# ----------------------------------------------------------------------

def main():
    """Main application entry point"""
    try:
        app = QApplication(sys.argv)
        
        # Set application-wide stylesheet
        app.setStyleSheet(MAIN_STYLESHEET)
        
        # Set application metadata
        app.setApplicationName("KTS Transport")
        app.setApplicationVersion("1.0")
        app.setOrganizationName("KTS Transport Company")
        
        # Initialize DB Manager
        db_manager = DBManager()

        # Show Login Dialog first
        login_dialog = LoginDialog(db_manager)
        if login_dialog.exec() == QDialog.DialogCode.Accepted:
            # If login is successful, create and show the main window
            window = MainWindow(db_manager)
            window.showMaximized()
            sys.exit(app.exec())
        else:
            # If login is cancelled or fails, exit the application
            sys.exit(0)
        
    except Exception as e:
        print(f"Application startup error: {e}")
        try:
            QMessageBox.critical(None, "Startup Error", f"Failed to start application: {str(e)}")
        except:
            print("Failed to show error message box")
        sys.exit(1)
        

if __name__ == "__main__":
    main()