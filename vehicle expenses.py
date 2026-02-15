import sys
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QLabel, QLineEdit,
    QDateEdit, QDialog, QFormLayout, QMessageBox, QComboBox, QHeaderView,
    QFileDialog
)
from PyQt6.QtCore import QDate, Qt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

blue_white_stylesheet = """
QMainWindow, QWidget, QDialog {
    background-color: #f0f5fa;
    color: #1a237e;
    font-family: Arial, sans-serif;
    font-size: 14px;
}
QLabel {
    color: #1a237e;
    font-weight: 600;
}
QLineEdit, QDateEdit, QComboBox {
    background-color: #ffffff;
    border: 2px solid #1a237e;
    border-radius: 5px;
    padding: 4px 8px;
    color: #1a237e;
}
QPushButton {
    background-color: #1a237e;
    color: #ffffff;
    border-radius: 6px;
    font-weight: bold;
    min-width: 55px;
    min-height: 28px;
}
QPushButton:hover {
    background-color: #3f51b5;
}
QTableWidget {
    background-color: #ffffff;
    gridline-color: #90caf9;
    color: #1a237e;
    border: none;
    font-size: 14px;
}
QHeaderView::section {
    background-color: #1a237e;
    color: white;
    font-weight: bold;
    font-size: 14px;
}
"""

class ExpenseDialog(QDialog):
    def __init__(self, parent=None, data=None):
        super().__init__(parent)
        self.setWindowTitle("New/Edit Expense Record")
        self.setMinimumWidth(420)
        self.setStyleSheet(blue_white_stylesheet)
        layout = QFormLayout(self)

        self.date_edit = QDateEdit(calendarPopup=True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setDate(QDate.currentDate())
        layout.addRow("Date:", self.date_edit)

        self.vehicle_no_edit = QLineEdit()
        layout.addRow("Vehicle No:", self.vehicle_no_edit)

        input_style = """
            QLineEdit {
                background: #fff;
                color: #1a237e;
                border: 2px solid #1a237e;
                border-radius: 5px;
                height: 28px;
                padding-left: 6px;
            }
            QLineEdit:focus {
                border-color: #2196f3;
            }
        """

        self.fc_expense_edit = QLineEdit()
        self.fc_expense_edit.setStyleSheet(input_style)
        layout.addRow("FC Expense:", self.fc_expense_edit)

        self.spare_work_amount_edit = QLineEdit()
        self.spare_work_amount_edit.setStyleSheet(input_style)
        self.spare_work_type_combo = QComboBox()
        self.spare_work_type_combo.addItems(
            ["None", "Spares", "Body Work", "Trailor Work", "Mechanical Work", "Electrical Work"]
        )
        self.spare_work_type_combo.setStyleSheet("""
            QComboBox {
                border: 2px solid #1a237e;
                border-radius: 5px;
                padding: 4px 8px;
                background-color: #fff;
                min-width: 140px;
                color: #1a237e;
            }
        """)
        spare_work_hbox = QHBoxLayout()
        spare_work_hbox.addWidget(self.spare_work_amount_edit)
        spare_work_hbox.addWidget(self.spare_work_type_combo)
        layout.addRow("Spare & Work:", spare_work_hbox)

        self.tyre_amount_edit = QLineEdit()
        self.tyre_amount_edit.setStyleSheet(input_style)
        self.tyre_type_combo = QComboBox()
        self.tyre_type_combo.addItems(["None", "New", "Old"])
        self.tyre_type_combo.setStyleSheet("""
            QComboBox {
                border: 2px solid #1a237e;
                border-radius: 5px;
                padding: 4px 8px;
                background-color: #fff;
                min-width: 70px;
                color: #1a237e;
            }
        """)
        tyre_hbox = QHBoxLayout()
        tyre_hbox.addWidget(self.tyre_amount_edit)
        tyre_hbox.addWidget(self.tyre_type_combo)
        layout.addRow("Tyre Expense:", tyre_hbox)

        self.tax_amount_edit = QLineEdit()
        self.tax_amount_edit.setStyleSheet(input_style)
        self.tax_type_combo = QComboBox()
        self.tax_type_combo.addItems(["None", "Quarter tax", "NP tax", "Green tax", "Five year tax"])
        self.tax_type_combo.setStyleSheet("""
            QComboBox {
                border: 2px solid #1a237e;
                border-radius: 5px;
                padding: 4px 8px;
                background-color: #fff;
                min-width: 90px;
                color: #1a237e;
            }
        """)
        tax_hbox = QHBoxLayout()
        tax_hbox.addWidget(self.tax_amount_edit)
        tax_hbox.addWidget(self.tax_type_combo)
        layout.addRow("Tax:", tax_hbox)

        self.loan_edit = QLineEdit()
        self.loan_edit.setStyleSheet(input_style)
        layout.addRow("Loan:", self.loan_edit)

        self.insurance_edit = QLineEdit()
        self.insurance_edit.setStyleSheet(input_style)
        layout.addRow("Insurance:", self.insurance_edit)

        self.others_edit = QLineEdit()
        self.others_edit.setStyleSheet(input_style)
        layout.addRow("Others:", self.others_edit)

        self.remarks_edit = QLineEdit()
        self.remarks_edit.setStyleSheet(input_style)
        layout.addRow("Remarks:", self.remarks_edit)

        self.save_btn = QPushButton("Save")
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #1a237e;
                color: #fff;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
                height: 32px;
            }
            QPushButton:hover {
                background-color: #3f51b5;
            }
        """)
        layout.addRow(self.save_btn)
        self.save_btn.clicked.connect(self.on_save_clicked)

        if data:
            data += ["0"] * (15 - len(data))
            self.date_edit.setDate(QDate.fromString(data[0], "yyyy-MM-dd"))
            self.vehicle_no_edit.setText(data[1])
            self.fc_expense_edit.setText(data[2])
            self.tyre_amount_edit.setText(data[3])
            self.tyre_type_combo.setCurrentText(data[4] if data[4] in ["New", "Old"] else "None")
            self.tax_amount_edit.setText(data[5])
            self.tax_type_combo.setCurrentText(data[6] if data[6] in ["Quarter tax", "NP tax", "Green tax", "Five year tax"] else "None")
            self.spare_work_amount_edit.setText(data[7])
            self.spare_work_type_combo.setCurrentText(data[8] if data[8] in ["None", "Spares", "Body Work", "Trailor Work", "Mechanical Work", "Electrical Work"] else "None")
            self.loan_edit.setText(data[9])
            self.insurance_edit.setText(data[10])
            self.others_edit.setText(data[11])
            self.remarks_edit.setText(data[12])

    def on_save_clicked(self):
        self.accept()

    def get_data(self):
        data = [
            self.date_edit.date().toString("yyyy-MM-dd"),
            self.vehicle_no_edit.text().strip(),
            self.fc_expense_edit.text().strip() or "0",
            self.tyre_amount_edit.text().strip() or "0",
            self.tyre_type_combo.currentText(),
            self.tax_amount_edit.text().strip() or "0",
            self.tax_type_combo.currentText(),
            self.spare_work_amount_edit.text().strip() or "0",
            self.spare_work_type_combo.currentText(),
            self.loan_edit.text().strip() or "0",
            self.insurance_edit.text().strip() or "0",
            self.others_edit.text().strip() or "0",
            self.remarks_edit.text().strip()
        ]
        return data

class VehicleExpensesWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Vehicle Expenses Manager")
        self.resize(1700, 900)
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)

        toolbar = QHBoxLayout()
        self.back_btn = QPushButton("BACK")
        self.back_btn.setStyleSheet("background-color: #555555; color: #fff; font-weight: bold; border-radius: 5px; min-width: 75px; height: 32px;")
        self.back_btn.clicked.connect(self.reset_filters)

        self.new_btn = QPushButton("NEW")
        self.new_btn.setStyleSheet("background-color: #1976d2; color: #fff; font-weight: bold; border-radius: 5px; min-width: 75px; height: 32px;")
        self.search_btn = QPushButton("SEARCH")
        self.search_btn.setStyleSheet("background-color: #d32f2f; color: #fff; font-weight: bold; border-radius: 5px; min-width: 75px; height: 32px;")
        self.reset_btn = QPushButton("RESET")
        self.reset_btn.setStyleSheet("background-color: #555555; color: #fff; font-weight: bold; border-radius: 5px; min-width: 75px; height: 32px;")
        self.download_btn = QPushButton("Download Excel")
        self.download_btn.setStyleSheet("background-color: #388e3c; color: #fff; font-weight: bold; border-radius: 5px; min-width: 120px; height: 32px;")
        self.download_btn.clicked.connect(self.download_to_excel)

        self.filter_vehicle = QLineEdit()
        self.filter_vehicle.setPlaceholderText("Vehicle No.")
        self.filter_vehicle.setStyleSheet("background: #fff; color: #1a237e; border-radius: 4px; border: 1px solid #1a237e; height: 28px;")
        self.date_filter_combo = QComboBox()
        self.date_filter_combo.addItems([
            "All",
            "Last 1 Day",
            "Last Month",
            "Last 3 Months",
            "Last 6 Months",
            "Last Year",
            "Custom",
        ])
        self.date_filter_combo.setCurrentIndex(0)
        self.date_filter_combo.setStyleSheet("background: #fff; color: #1a237e; border-radius: 4px; border: 1px solid #1a237e; height: 28px;")
        self.date_filter_combo.currentTextChanged.connect(self.on_date_filter_changed)
        self.custom_from_date = QDateEdit(calendarPopup=True)
        self.custom_from_date.setDisplayFormat("yyyy-MM-dd")
        self.custom_from_date.setDate(QDate.currentDate().addMonths(-1))
        self.custom_from_date.setStyleSheet("background: #fff; color: #1a237e; border-radius: 4px; border: 1px solid #1a237e; height: 28px;")
        self.custom_from_date.setVisible(False)
        self.custom_to_date = QDateEdit(calendarPopup=True)
        self.custom_to_date.setDisplayFormat("yyyy-MM-dd")
        self.custom_to_date.setDate(QDate.currentDate())
        self.custom_to_date.setStyleSheet("background: #fff; color: #1a237e; border-radius: 4px; border: 1px solid #1a237e; height: 28px;")
        self.custom_to_date.setVisible(False)

        toolbar.addWidget(self.back_btn)
        toolbar.addWidget(self.new_btn)
        toolbar.addWidget(QLabel("Vehicle:"))
        toolbar.addWidget(self.filter_vehicle)
        toolbar.addWidget(QLabel("Date Filter:"))
        toolbar.addWidget(self.date_filter_combo)
        toolbar.addWidget(QLabel("From:"))
        toolbar.addWidget(self.custom_from_date)
        toolbar.addWidget(QLabel("To:"))
        toolbar.addWidget(self.custom_to_date)
        toolbar.addWidget(self.search_btn)
        toolbar.addWidget(self.reset_btn)
        toolbar.addWidget(self.download_btn)
        toolbar.addStretch()
        main_layout.addLayout(toolbar)

        self.table = QTableWidget(0, 16)
        header_names = [
            "Date", "Vehicle No", "FC Expense", "Tyre Amount", "Tyre Type",
            "Tax", "Tax Type", "Spare & Work", "Type",
            "Loan", "Insurance", "Others", "Remarks", "Total", "Edit", "Delete"
        ]
        self.table.setHorizontalHeaderLabels(header_names)
        self.table.setStyleSheet(blue_white_stylesheet)

        header = self.table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(header_names)):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            self.table.horizontalHeaderItem(i).setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        main_layout.addWidget(self.table)

        self.grand_total_label = QLabel("Grand Total: 0.00")
        self.grand_total_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #1a237e; padding: 8px;")
        main_layout.addWidget(self.grand_total_label)

        self.records = []
        self.new_btn.clicked.connect(self.open_new_dialog)
        self.search_btn.clicked.connect(self.apply_filters)
        self.reset_btn.clicked.connect(self.reset_filters)

    def on_date_filter_changed(self, text):
        is_custom = (text == "Custom")
        self.custom_from_date.setVisible(is_custom)
        self.custom_to_date.setVisible(is_custom)

    def open_new_dialog(self):
        dialog = ExpenseDialog(self)
        if dialog.exec():
            data = dialog.get_data()
            data += ["0"] * (16 - len(data))  # Pad to 16 for total, edit, delete columns
            self.records.insert(0, data)
            self.refresh_table()

    def open_edit_dialog(self, index):
        data = self.records[index]
        data += ["0"] * (16 - len(data))
        dialog = ExpenseDialog(self, data)
        if dialog.exec():
            self.records[index] = dialog.get_data()
            self.refresh_table()

    def delete_record(self, index):
        reply = QMessageBox.question(
            self, "Confirm Delete", "Are you sure you want to delete this record?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            del self.records[index]
            self.refresh_table()

    def refresh_table(self, filtered_records=None):
        rows = filtered_records if filtered_records is not None else self.records
        self.table.setRowCount(len(rows))
        grand_total = 0.0

        for row, record in enumerate(rows):
            for col in range(13):
                value = record[col] if col < len(record) else ""
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, col, item)

            total = 0.0
            for idx in [2, 3, 5, 7, 9, 10, 11]:
                try:
                    total += float(record[idx])
                except (ValueError, IndexError):
                    pass

            total_item = QTableWidgetItem(f"{total:.2f}")
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            total_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
            self.table.setItem(row, 13, total_item)
            grand_total += total

            edit_btn = QPushButton("Edit")
            edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            edit_btn.setStyleSheet("min-width: 50px; min-height: 22px; font-size: 13px; background-color: #1976d2; color: #fff; border-radius: 4px; font-weight: bold;")
            record_index = self.records.index(record)
            edit_btn.clicked.connect(lambda _, ix=record_index: self.open_edit_dialog(ix))
            self.table.setCellWidget(row, 14, edit_btn)

            del_btn = QPushButton("Delete")
            del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            del_btn.setStyleSheet("min-width: 50px; min-height: 22px; font-size: 13px; background-color: #d32f2f; color: #fff; border-radius: 4px; font-weight: bold;")
            del_btn.clicked.connect(lambda _, ix=record_index: self.delete_record(ix))
            self.table.setCellWidget(row, 15, del_btn)

        self.grand_total_label.setText(f"Grand Total: {grand_total:.2f}")

    def apply_filters(self):
        vehicle_filter = self.filter_vehicle.text().strip().lower()
        date_filter = self.date_filter_combo.currentText()

        today = QDate.currentDate()
        date_from, date_to = None, today
        if date_filter == "All":
            date_from, date_to = None, None
        elif date_filter == "Last 1 Day":
            date_from = today.addDays(-1)
        elif date_filter == "Last Month":
            date_from = today.addMonths(-1)
        elif date_filter == "Last 3 Months":
            date_from = today.addMonths(-3)
        elif date_filter == "Last 6 Months":
            date_from = today.addMonths(-6)
        elif date_filter == "Last Year":
            date_from = today.addYears(-1)
        elif date_filter == "Custom":
            date_from = self.custom_from_date.date()
            date_to = self.custom_to_date.date()
            if date_to < date_from:
                QMessageBox.warning(self, "Invalid Date Range", "'To' date must be after or equal to 'From' date.")
                return

        filtered = []
        for record in self.records:
            vehicle_match = (not vehicle_filter) or (len(record) > 1 and record[1].lower() == vehicle_filter)
            try:
                rec_date = QDate.fromString(record[0], "yyyy-MM-dd")
            except Exception:
                rec_date = None
            date_match = True
            if date_from and date_to and rec_date:
                date_match = date_from <= rec_date <= date_to
            if vehicle_match and date_match:
                filtered.append(record)
        self.refresh_table(filtered)

    def reset_filters(self):
        self.filter_vehicle.clear()
        self.date_filter_combo.setCurrentIndex(0)
        self.custom_from_date.setDate(QDate.currentDate().addMonths(-1))
        self.custom_to_date.setDate(QDate.currentDate())
        self.custom_from_date.setVisible(False)
        self.custom_to_date.setVisible(False)
        self.refresh_table()

    def download_to_excel(self):
        if not self.records:
            QMessageBox.information(self, "No Data", "No records available to download.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Expenses"

        headers = [
            "Date", "Vehicle No", "FC Expense", "Tyre Amount", "Tyre Type",
            "Tax", "Tax Type", "Spare & Work", "Type",
            "Loan", "Insurance", "Others", "Remarks", "Total"
        ]

        font_bold = Font(bold=True)
        align_center = Alignment(horizontal='center')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = font_bold
            cell.alignment = align_center

        numeric_col_indices = [3, 4, 6, 8, 10, 11, 14] # corresponding to FC Expense, Tyre Amount, Tax, Spare & Work Amount, Loan, Insurance, Others, Total

        totals = [0] * len(headers)
        for row_num, record in enumerate(self.records, 2):
            row_total = 0.0
            for col_num in range(len(headers) - 1):  # excluding Total column
                val = record[col_num] if col_num < len(record) else ""
                ws.cell(row=row_num, column=col_num + 1, value=val)
                if (col_num + 1) in numeric_col_indices[:-1]:  # exclude Total column here
                    try:
                        fval = float(val)
                        totals[col_num] += fval
                        row_total += fval
                    except:
                        pass
            ws.cell(row=row_num, column=len(headers), value=row_total)
            totals[len(headers)-1] += row_total

        total_row = len(self.records) + 2
        ws.cell(row=total_row, column=1, value="Totals:")
        ws.cell(row=total_row, column=1).font = font_bold
        ws.cell(row=total_row, column=1).alignment = align_center
        for col in range(2, len(headers) + 1):
            if col in numeric_col_indices:
                val = totals[col-1]
                ws.cell(row=total_row, column=col, value=val)
                ws.cell(row=total_row, column=col).font = font_bold
                ws.cell(row=total_row, column=col).alignment = align_center
            else:
                ws.cell(row=total_row, column=col, value="")
                ws.cell(row=total_row, column=col).font = font_bold
                ws.cell(row=total_row, column=col).alignment = align_center

        try:
            wb.save(path)
            QMessageBox.information(self, "Success", f"Records successfully saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(blue_white_stylesheet)
    window = VehicleExpensesWindow()
    window.show()
    sys.exit(app.exec())
